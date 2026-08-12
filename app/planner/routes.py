import os
from datetime import date, datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

import calendar
import io
import json
from flask import render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy.orm.attributes import flag_modified
from app import db
from app.planner import planner
from app.models import DailyPlan, MonthlyPlan, YearlyPlan, WeeklyPlan, User, PlanningTask
from app.services.cascade_service import (
    get_yearly_events_for_month,
    get_monthly_items_for_date,
    get_weekly_todos_for_date,
    get_all_cascaded_items_for_daily
)

DEFAULT_TAGS = [
    {'id': 'tag_work', 'name': 'Work', 'color': '#3b82f6'},
    {'id': 'tag_personal', 'name': 'Personal', 'color': '#10b981'},
    {'id': 'tag_health', 'name': 'Health', 'color': '#ec4899'},
    {'id': 'tag_finance', 'name': 'Finance', 'color': '#f59e0b'},
    {'id': 'tag_urgent', 'name': 'Urgent', 'color': '#ef4444'}
]

def get_user_tags(user):
    if user and hasattr(user, 'custom_tags') and user.custom_tags:
        return user.custom_tags
    return DEFAULT_TAGS

def get_today_date():
    """Return current date in configured timezone (default Asia/Kolkata / IST UTC+5:30)."""
    tz_name = os.environ.get('APP_TIMEZONE', 'Asia/Kolkata')
    if ZoneInfo:
        try:
            return datetime.now(ZoneInfo(tz_name)).date()
        except Exception:
            pass
    return datetime.now(timezone(timedelta(hours=5, minutes=30))).date()

@planner.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('planner.dashboard'))
    return redirect(url_for('auth.login'))


@planner.before_app_request
def auto_daily_drive_sync():
    """Triggers automatic Google Drive backup once per day when authenticated user accesses the app."""
    if current_user and current_user.is_authenticated:
        from app.services.google_service import check_and_trigger_daily_drive_sync
        res = check_and_trigger_daily_drive_sync(current_user)
        if res and isinstance(res, dict) and res.get('success'):
            flash('Daily automatic Google Drive backup completed!', 'success')


def process_task_spillovers(user_id, target_date):
    """
    Automatically rolls over uncompleted non-default tasks from past dates up to target_date.
    Increments spillover_count (pending days) and escalates priority up to 'High' (max severity).
    Restricted to past 30 days for maximum performance.
    """
    cutoff_date = target_date - timedelta(days=30)
    past_plans = DailyPlan.query.filter(
        DailyPlan.user_id == user_id,
        DailyPlan.date < target_date,
        DailyPlan.date >= cutoff_date
    ).order_by(DailyPlan.date.asc()).all()

    if not past_plans:
        return

    # Quick check: if no past plan contains uncompleted non-default tasks, return immediately!
    has_uncompleted = any(
        any(isinstance(t, dict) and not t.get('completed') and not t.get('is_default') for t in (p.tasks or []))
        for p in past_plans
    )
    if not has_uncompleted:
        return

    plan_map = {p.date: p for p in past_plans}
    earliest_date = past_plans[0].date

    curr_date = earliest_date
    while curr_date < target_date:
        next_date = curr_date + timedelta(days=1)
        curr_plan = plan_map.get(curr_date)

        if curr_plan and curr_plan.tasks:
            # Only non-default uncompleted tasks spill over
            uncompleted = [t for t in curr_plan.tasks if isinstance(t, dict) and not t.get('completed') and not t.get('is_default')]
            if uncompleted:
                next_plan = plan_map.get(next_date)
                if not next_plan:
                    next_plan = DailyPlan.query.filter_by(user_id=user_id, date=next_date).first()
                if not next_plan:
                    next_plan = DailyPlan(user_id=user_id, date=next_date, schedule={}, tasks=[], notes='')
                    db.session.add(next_plan)

                next_tasks = list(next_plan.tasks or [])
                existing_ids = {t.get('id') for t in next_tasks if isinstance(t, dict) and t.get('id')}

                dismissed_ids = set()
                raw_sched = next_plan.schedule or {}
                if isinstance(raw_sched, dict):
                    dismissed_ids = set(raw_sched.get('_dismissed_tasks', []))

                modified_next = False

                for t in uncompleted:
                    t_id = t.get('id')
                    if t_id and t_id not in existing_ids and t_id not in dismissed_ids:
                        old_priority = t.get('priority', 'Medium')
                        if old_priority == 'Low':
                            new_priority = 'Medium'
                        else:
                            new_priority = 'High'  # Escalates to High (max severity)

                        spill_count = t.get('spillover_count', 0) + 1
                        orig_date = t.get('original_date', curr_date.strftime('%Y-%m-%d'))

                        spill_task = {
                            'id': t_id,
                            'text': t.get('text', ''),
                            'priority': new_priority,
                            'completed': False,
                            'is_spillover': True,
                            'spillover_count': spill_count,
                            'original_date': orig_date
                        }
                        next_tasks.append(spill_task)
                        existing_ids.add(t_id)
                        modified_next = True

                if modified_next:
                    next_plan.tasks = next_tasks
                    flag_modified(next_plan, 'tasks')
                    plan_map[next_date] = next_plan

        curr_date = next_date

    db.session.commit()


def populate_daily_defaults(user_id, target_date):
    """
    Populates default daily routine tasks and default activity schedule slots
    for target_date based on the latest saved default configurations.
    """
    past_plans = DailyPlan.query.filter(
        DailyPlan.user_id == user_id,
        DailyPlan.date <= target_date
    ).order_by(DailyPlan.date.desc()).limit(30).all()

    if not past_plans:
        return

    # Resolve active Default Tasks based on the MOST RECENT plan setting/action for each task
    default_tasks_master = {}
    seen_task_keys = set()

    for p in past_plans:
        raw_sched = p.schedule or {}
        deleted_on_plan = set()
        if isinstance(raw_sched, dict):
            for dkey in raw_sched.get('_deleted_defaults', []):
                deleted_on_plan.add(str(dkey).lower())

        if p.tasks:
            for t in p.tasks:
                if isinstance(t, dict):
                    t_text = t.get('text', '').strip()
                    t_id = str(t.get('id', '')).lower()
                    t_key = t_text.lower()

                    if t_key and t_key not in seen_task_keys:
                        seen_task_keys.add(t_key)
                        # If marked default and not deleted on this same plan
                        if t.get('is_default') and t_key not in deleted_on_plan and t_id not in deleted_on_plan:
                            default_tasks_master[t_key] = {
                                'text': t_text,
                                'priority': t.get('priority', 'Medium'),
                                'is_default': True
                            }

        # Any default deletion recorded on this plan that wasn't already determined by a newer plan
        for dkey in deleted_on_plan:
            seen_task_keys.add(dkey)

    # Resolve Default Schedule Slots based on the MOST RECENT plan setting for each slot
    default_schedule_master = {}
    resolved_slots = set()

    for p in past_plans:
        if p.schedule:
            for slot, sdata in p.schedule.items():
                if slot.startswith('_') or slot in resolved_slots:
                    continue

                if isinstance(sdata, dict):
                    if p.date < target_date or sdata.get('activity') or sdata.get('is_default'):
                        resolved_slots.add(slot)
                        if sdata.get('is_default') and sdata.get('activity'):
                            default_schedule_master[slot] = {
                                'activity': sdata.get('activity', ''),
                                'mood': sdata.get('mood', ''),
                                'is_default': True
                            }

    if not default_tasks_master and not default_schedule_master:
        return

    plan = DailyPlan.query.filter_by(user_id=user_id, date=target_date).first()
    if not plan:
        plan = DailyPlan(user_id=user_id, date=target_date, schedule={}, tasks=[], notes='')
        db.session.add(plan)

    tasks = list(plan.tasks or [])
    schedule = dict(plan.schedule or {})
    modified = False

    target_dismissed = set()
    if isinstance(schedule, dict):
        target_dismissed = set(schedule.get('_dismissed_tasks', []))
        target_dismissed.update(schedule.get('_deleted_defaults', []))

    target_dismissed_lower = {str(x).lower() for x in target_dismissed}

    # Populate missing default tasks
    existing_texts = {t.get('text', '').strip().lower() for t in tasks if isinstance(t, dict) and t.get('text')}
    for key_lower, dtask in default_tasks_master.items():
        if key_lower not in existing_texts and key_lower not in target_dismissed_lower:
            new_id = f"def_{int(datetime.utcnow().timestamp() * 1000)}_{len(tasks)}"
            if new_id.lower() not in target_dismissed_lower:
                tasks.append({
                    'id': new_id,
                    'text': dtask['text'],
                    'priority': dtask['priority'],
                    'completed': False,
                    'is_default': True,
                    'is_spillover': False,
                    'spillover_count': 0,
                    'original_date': target_date.strftime('%Y-%m-%d')
                })
                modified = True

    # Populate missing default schedule slots
    for slot, dslot in default_schedule_master.items():
        curr_val = schedule.get(slot)
        is_empty = False
        if not curr_val:
            is_empty = True
        elif isinstance(curr_val, dict) and not curr_val.get('activity'):
            is_empty = True

        if is_empty:
            schedule[slot] = {
                'activity': dslot['activity'],
                'mood': dslot['mood'],
                'is_default': True
            }
            modified = True

    if modified:
        plan.tasks = tasks
        plan.schedule = schedule
        flag_modified(plan, 'tasks')
        flag_modified(plan, 'schedule')
        db.session.commit()


@planner.route('/dashboard')
@login_required
def dashboard():
    today = get_today_date()
    current_year = today.year
    current_month = today.month

    # Process automatic task spillovers and daily routine defaults up to today
    process_task_spillovers(current_user.id, today)
    populate_daily_defaults(current_user.id, today)

    # Get or create today's daily plan
    daily_plan = DailyPlan.query.filter_by(user_id=current_user.id, date=today).first()
    today_tasks = list(daily_plan.tasks or []) if daily_plan else []
    today_tasks.sort(key=lambda t: (1 if t.get('completed') else 0, 0 if t.get('priority') == 'High' else (1 if t.get('priority') == 'Medium' else 2)))

    completed_today = sum(1 for t in today_tasks if t.get('completed'))
    total_today = len(today_tasks)
    today_completion_pct = int((completed_today / total_today * 100)) if total_today > 0 else 0

    # Get current month's plan
    monthly_plan = MonthlyPlan.query.filter_by(user_id=current_user.id, year=current_year, month=current_month).first()
    monthly_goals = monthly_plan.goals if monthly_plan and monthly_plan.goals else []
    monthly_habits = monthly_plan.habits if monthly_plan and monthly_plan.habits else []
    goals_completed = sum(1 for g in monthly_goals if g.get('status') == 'Completed')

    # Get current year's plan
    yearly_plan = YearlyPlan.query.filter_by(user_id=current_user.id, year=current_year).first()
    yearly_resolutions = yearly_plan.resolutions if yearly_plan and yearly_plan.resolutions else []
    yearly_objectives = yearly_plan.objectives if yearly_plan and yearly_plan.objectives else []

    # Format month name
    month_name = calendar.month_name[current_month]

    # Aggregate Depression Tracker Analytics across recent DailyPlan entries
    all_daily_plans = DailyPlan.query.filter_by(user_id=current_user.id).order_by(DailyPlan.date.desc()).limit(30).all()
    recent_episodes = []
    total_episodes_count = 0
    intensity_sum = 0
    peak_intensity = 0
    coping_strategies_map = {}

    for dp in all_daily_plans:
        if dp.depression_episodes:
            for ep in dp.depression_episodes:
                total_episodes_count += 1
                intensity = int(ep.get('intensity', 5))
                intensity_sum += intensity
                if intensity > peak_intensity:
                    peak_intensity = intensity

                coping = ep.get('coping_mechanism', '').strip()
                if coping:
                    coping_strategies_map[coping] = coping_strategies_map.get(coping, 0) + 1

                ep_copy = dict(ep)
                ep_copy['date_str'] = dp.date.strftime('%b %d, %Y')
                recent_episodes.append(ep_copy)

    avg_intensity = round(intensity_sum / total_episodes_count, 1) if total_episodes_count > 0 else 0
    top_coping_strategies = sorted(coping_strategies_map.items(), key=lambda x: x[1], reverse=True)[:3]

    # Aggregate Memory Tracker Analytics across recent DailyPlan entries
    recent_memory_logs = []
    total_memory_slips_count = 0
    recovered_slips_count = 0
    memory_category_map = {}

    for dp in all_daily_plans:
        if dp.memory_logs:
            for log in dp.memory_logs:
                total_memory_slips_count += 1
                recovery = log.get('recovery', '').strip()
                if recovery and 'unresolved' not in recovery.lower() and 'still forgot' not in recovery.lower():
                    recovered_slips_count += 1
                
                cat = log.get('category', 'General').strip()
                if cat:
                    memory_category_map[cat] = memory_category_map.get(cat, 0) + 1
                
                log_copy = dict(log)
                log_copy['date_str'] = dp.date.strftime('%b %d, %Y')
                recent_memory_logs.append(log_copy)

    memory_recovery_pct = int((recovered_slips_count / total_memory_slips_count) * 100) if total_memory_slips_count > 0 else 0
    top_memory_categories = sorted(memory_category_map.items(), key=lambda x: x[1], reverse=True)[:3]
    today_memory_count = len(daily_plan.memory_logs or []) if daily_plan and daily_plan.memory_logs else 0

    # Aggregate Sleep Tracker Analytics across recent DailyPlan entries
    recent_sleep_logs = []
    total_sleep_hours_sum = 0
    sleep_days_count = 0
    sleep_quality_sum = 0

    for dp in all_daily_plans:
        if dp.sleep_log and isinstance(dp.sleep_log, dict) and dp.sleep_log.get('hours'):
            try:
                s_hours = float(dp.sleep_log.get('hours', 0))
            except (ValueError, TypeError):
                s_hours = 0
            try:
                s_quality = int(dp.sleep_log.get('quality', 0))
            except (ValueError, TypeError):
                s_quality = 0

            if s_hours > 0:
                total_sleep_hours_sum += s_hours
                sleep_days_count += 1
                sleep_quality_sum += s_quality

                log_copy = dict(dp.sleep_log)
                log_copy['date_str'] = dp.date.strftime('%b %d, %Y')
                recent_sleep_logs.append(log_copy)

    avg_sleep_hours = round(total_sleep_hours_sum / sleep_days_count, 1) if sleep_days_count > 0 else 0
    avg_sleep_quality = round(sleep_quality_sum / sleep_days_count, 1) if sleep_days_count > 0 else 0

    # 5. Dashboard Reminder Panel (Marquee Alerts)
    marquee_alerts = []

    # 1) Unchecked Shopping List Items (Pending for X days)
    recent_weekly_plans = WeeklyPlan.query.filter_by(user_id=current_user.id).order_by(WeeklyPlan.start_date.desc()).limit(8).all()
    for wp in recent_weekly_plans:
        if wp.shopping_list:
            for item in wp.shopping_list:
                if not item.get('bought'):
                    added_str = item.get('added_date')
                    item_name = item.get('item', 'Item')
                    if added_str:
                        try:
                            a_date = datetime.strptime(added_str, '%Y-%m-%d').date()
                            days_pending = (today - a_date).days
                        except ValueError:
                            days_pending = 0
                    else:
                        days_pending = 0

                    if days_pending < 0:
                        days_pending = 0

                    marquee_alerts.append({
                        'type': 'shopping',
                        'icon': 'fa-solid fa-cart-shopping',
                        'badge': '🛒 Shopping Alert',
                        'badge_color': 'bg-amber-500/20 text-amber-300 border-amber-500/30',
                        'text': f'"{item_name}" is pending in your Shopping List for {days_pending} day{"s" if days_pending != 1 else ""}'
                    })

    # 2) Today's Birthdays and Anniversaries
    if yearly_plan and yearly_plan.events:
        for ev in yearly_plan.events:
            ev_date_str = ev.get('date', '')
            ev_type = ev.get('event_type', 'event')
            title = ev.get('title', 'Event')
            
            is_today = False
            if ev_date_str:
                try:
                    ev_date = datetime.strptime(ev_date_str, '%Y-%m-%d').date()
                    if ev_date.month == today.month and ev_date.day == today.day:
                        is_today = True
                except ValueError:
                    try:
                        parts = [p.strip() for p in ev_date_str.split('-') if p.strip()]
                        if len(parts) == 3 and int(parts[1]) == today.month and int(parts[2]) == today.day:
                            is_today = True
                        elif len(parts) == 2 and int(parts[0]) == today.month and int(parts[1]) == today.day:
                            is_today = True
                    except Exception:
                        pass
                
                if is_today:
                    if ev_type == 'birthday':
                        marquee_alerts.append({
                            'type': 'birthday',
                            'icon': 'fa-solid fa-cake-candles',
                            'badge': '🎂 Birthday Today',
                            'badge_color': 'bg-rose-500/20 text-rose-300 border-rose-500/30',
                            'text': f'Wish {title} a Happy Birthday today!'
                        })
                    elif ev_type == 'anniversary':
                        marquee_alerts.append({
                            'type': 'anniversary',
                            'icon': 'fa-solid fa-ring',
                            'badge': '💍 Anniversary Today',
                            'badge_color': 'bg-purple-500/20 text-purple-300 border-purple-500/30',
                            'text': f'Celebration: {title} is today!'
                        })
                    else:
                        marquee_alerts.append({
                            'type': 'annual_event',
                            'icon': 'fa-solid fa-calendar-day',
                            'badge': 'Annual Event',
                            'badge_color': 'bg-emerald-500/20 text-emerald-300 border-emerald-500/30',
                            'text': f'Annual Event Today: "{title}"'
                        })

    # 3) Today's High Priority Events / Tasks / Deadlines
    for t in today_tasks:
        if t.get('priority') == 'High' and not t.get('completed'):
            marquee_alerts.append({
                'type': 'high_priority',
                'icon': 'fa-solid fa-fire',
                'badge': '🔥 High Priority',
                'badge_color': 'bg-rose-500/20 text-rose-300 border-rose-500/40',
                'text': f'High priority task for today: "{t.get("text")}"'
            })

    if monthly_plan and monthly_plan.calendar_days:
        today_day_str = str(today.day)
        for d_key, d_data in monthly_plan.calendar_days.items():
            if str(d_key).lstrip('0') == today_day_str:
                if isinstance(d_data, dict):
                    for citem in d_data.get('items', []):
                        is_remind = citem.get('remind_me') in [True, 'true', 'True', '1', 1, 'on']
                        if is_remind:
                            marquee_alerts.append({
                                'type': 'reminder',
                                'icon': 'fa-solid fa-bell',
                                'badge': 'Remind Me',
                                'badge_color': 'bg-amber-500/20 text-amber-300 border-amber-500/30',
                                'text': f'Reminder for today: "{citem.get("text")}"'
                            })
                        elif citem.get('type') == 'deadline' or 'interview' in citem.get('text', '').lower() or 'deadline' in citem.get('text', '').lower():
                            marquee_alerts.append({
                                'type': 'deadline',
                                'icon': 'fa-solid fa-calendar-check',
                                'badge': '⏰ Today\'s Deadline',
                                'badge_color': 'bg-indigo-500/20 text-indigo-300 border-indigo-500/30',
                                'text': f'Deadline / Event scheduled today: "{citem.get("text")}"'
                            })

    return render_template(
        'planner/dashboard.html',
        today=today,
        today_tasks=today_tasks,
        completed_today=completed_today,
        total_today=total_today,
        today_completion_pct=today_completion_pct,
        daily_plan=daily_plan,
        monthly_plan=monthly_plan,
        monthly_goals=monthly_goals,
        monthly_habits=monthly_habits,
        goals_completed=goals_completed,
        yearly_plan=yearly_plan,
        yearly_resolutions=yearly_resolutions,
        yearly_objectives=yearly_objectives,
        current_year=current_year,
        month_name=month_name,
        recent_episodes=recent_episodes[:5],
        total_episodes_count=total_episodes_count,
        avg_intensity=avg_intensity,
        peak_intensity=peak_intensity,
        top_coping_strategies=top_coping_strategies,
        recent_memory_logs=recent_memory_logs[:5],
        total_memory_slips_count=total_memory_slips_count,
        memory_recovery_pct=memory_recovery_pct,
        top_memory_categories=top_memory_categories,
        today_memory_count=today_memory_count,
        recent_sleep_logs=recent_sleep_logs[:5],
        avg_sleep_hours=avg_sleep_hours,
        avg_sleep_quality=avg_sleep_quality,
        sleep_days_count=sleep_days_count,
        marquee_alerts=marquee_alerts
    )


def carry_forward_unbought_shopping_items(user_id, target_year, target_week):
    """
    Moves unbought shopping list items (bought: False) from prior weekly plans
    to the target week's shopping list, and removes them from prior weekly plans.
    """
    first_day_of_year = date(target_year, 1, 4)
    start_of_target_week = first_day_of_year + timedelta(weeks=target_week - 1) - timedelta(days=first_day_of_year.weekday())

    target_plan = WeeklyPlan.query.filter_by(
        user_id=user_id,
        year=target_year,
        week_number=target_week
    ).first()

    prior_plans = WeeklyPlan.query.filter(
        WeeklyPlan.user_id == user_id,
        WeeklyPlan.start_date < start_of_target_week
    ).order_by(WeeklyPlan.start_date.asc()).all()

    items_to_move = []
    modified_priors = False

    for prior_plan in prior_plans:
        p_list = prior_plan.shopping_list or []
        if not p_list:
            continue
        
        bought_items = []
        unbought_items = []

        for item in p_list:
            if isinstance(item, dict):
                if item.get('bought'):
                    bought_items.append(item)
                else:
                    unbought_items.append(item)

        if unbought_items:
            items_to_move.extend(unbought_items)
            prior_plan.shopping_list = bought_items
            flag_modified(prior_plan, 'shopping_list')
            modified_priors = True

    if items_to_move:
        if not target_plan:
            target_plan = WeeklyPlan(
                user_id=user_id,
                year=target_year,
                week_number=target_week,
                start_date=start_of_target_week,
                goals=[],
                daily_todos={abbr: [] for abbr in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']},
                shopping_list=[],
                meals_menu={abbr: {'breakfast': '', 'lunch': '', 'dinner': ''} for abbr in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']},
                notes=''
            )
            db.session.add(target_plan)

        target_list = list(target_plan.shopping_list or [])
        existing_ids = {s.get('id') for s in target_list if isinstance(s, dict) and s.get('id')}

        for item in items_to_move:
            if item.get('id') not in existing_ids:
                target_list.append(item)

        target_plan.shopping_list = target_list
        flag_modified(target_plan, 'shopping_list')
        db.session.commit()
    elif modified_priors:
        db.session.commit()

    return target_plan


@planner.route('/weekly', methods=['GET', 'POST'])
@login_required
def weekly():
    today = get_today_date()
    current_year, current_week, _ = today.isocalendar()

    year_param = request.args.get('year', type=int)
    week_param = request.args.get('week', type=int)

    if not year_param or not week_param:
        year_param, week_param = current_year, current_week

    first_day_of_year = date(year_param, 1, 4)
    start_of_week = first_day_of_year + timedelta(weeks=week_param - 1) - timedelta(days=first_day_of_year.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    day_abbrs = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    days_of_week = []
    
    # Automatically move unbought shopping list items from prior weeks to target week
    plan = carry_forward_unbought_shopping_items(current_user.id, year_param, week_param)
    if not plan:
        plan = WeeklyPlan.query.filter_by(user_id=current_user.id, year=year_param, week_number=week_param).first()

    if request.method == 'POST':
        action = request.form.get('action')

        if not plan:
            plan = WeeklyPlan(
                user_id=current_user.id,
                year=year_param,
                week_number=week_param,
                start_date=start_of_week,
                goals=[],
                daily_todos={abbr: [] for abbr in day_abbrs},
                shopping_list=[],
                meals_menu={abbr: {'breakfast': '', 'lunch': '', 'dinner': ''} for abbr in day_abbrs},
                notes=''
            )
            db.session.add(plan)

        goals = plan.goals or []
        daily_todos = plan.daily_todos or {abbr: [] for abbr in day_abbrs}
        shopping_list = plan.shopping_list or []
        meals_menu = plan.meals_menu or {abbr: {'breakfast': '', 'lunch': '', 'dinner': ''} for abbr in day_abbrs}

        if action == 'add_weekly_goal':
            goal_title = request.form.get('goal_title', '').strip()
            if goal_title:
                goals.append({
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'title': goal_title,
                    'completed': False
                })
                plan.goals = goals
                flag_modified(plan, 'goals')
                db.session.commit()
                flash('Weekly goal added!', 'success')

        elif action == 'toggle_weekly_goal':
            goal_id = request.form.get('goal_id')
            for g in goals:
                if g.get('id') == goal_id:
                    g['completed'] = not g.get('completed', False)
            plan.goals = goals
            flag_modified(plan, 'goals')
            db.session.commit()

        elif action == 'delete_weekly_goal':
            goal_id = request.form.get('goal_id')
            plan.goals = [g for g in goals if g.get('id') != goal_id]
            flag_modified(plan, 'goals')
            db.session.commit()
            flash('Weekly goal removed.', 'info')

        elif action == 'add_daily_todo':
            day_abbr = request.form.get('day_abbr')
            todo_text = request.form.get('todo_text', '').strip()
            if day_abbr in day_abbrs and todo_text:
                if day_abbr not in daily_todos:
                    daily_todos[day_abbr] = []
                daily_todos[day_abbr].append({
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'text': todo_text,
                    'completed': False
                })
                plan.daily_todos = daily_todos
                flag_modified(plan, 'daily_todos')
                db.session.commit()
                flash(f'To-do added to {day_abbr}!', 'success')

        elif action == 'toggle_daily_todo':
            day_abbr = request.form.get('day_abbr')
            todo_id = request.form.get('todo_id')
            if day_abbr in daily_todos:
                for t in daily_todos[day_abbr]:
                    if t.get('id') == todo_id:
                        t['completed'] = not t.get('completed', False)
                plan.daily_todos = daily_todos
                flag_modified(plan, 'daily_todos')
                db.session.commit()

        elif action == 'delete_daily_todo':
            day_abbr = request.form.get('day_abbr')
            todo_id = request.form.get('todo_id')
            if day_abbr in daily_todos:
                daily_todos[day_abbr] = [t for t in daily_todos[day_abbr] if t.get('id') != todo_id]
                plan.daily_todos = daily_todos
                flag_modified(plan, 'daily_todos')
                db.session.commit()
                flash('To-do item deleted.', 'info')

        elif action == 'add_shopping_item':
            item_name = request.form.get('item_name', '').strip()
            category = request.form.get('category', 'Groceries').strip()
            if item_name:
                shopping_list.append({
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'item': item_name,
                    'category': category,
                    'bought': False,
                    'added_date': get_today_date().strftime('%Y-%m-%d')
                })
                plan.shopping_list = shopping_list
                flag_modified(plan, 'shopping_list')
                db.session.commit()
                flash('Shopping item added!', 'success')

        elif action == 'toggle_shopping_item':
            item_id = request.form.get('item_id')
            for s in shopping_list:
                if s.get('id') == item_id:
                    s['bought'] = not s.get('bought', False)
            plan.shopping_list = shopping_list
            flag_modified(plan, 'shopping_list')
            db.session.commit()

        elif action == 'delete_shopping_item':
            item_id = request.form.get('item_id')
            plan.shopping_list = [s for s in shopping_list if s.get('id') != item_id]
            flag_modified(plan, 'shopping_list')
            db.session.commit()
            flash('Shopping item deleted.', 'info')

        elif action == 'save_meals_menu':
            for abbr in day_abbrs:
                meals_menu[abbr] = {
                    'breakfast': request.form.get(f'meal_bf_{abbr}', '').strip(),
                    'lunch': request.form.get(f'meal_lu_{abbr}', '').strip(),
                    'dinner': request.form.get(f'meal_dn_{abbr}', '').strip()
                }
            plan.meals_menu = meals_menu
            flag_modified(plan, 'meals_menu')
            db.session.commit()
            flash('Weekly meal menu saved!', 'success')

        elif action == 'save_weekly_notes':
            notes = request.form.get('notes', '').strip()
            plan.notes = notes
            db.session.commit()
            flash('Weekly notes updated!', 'success')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('is_ajax') == 'true':
            tot_todos = sum(len(daily_todos.get(abbr, [])) for abbr in day_abbrs)
            comp_todos = sum(sum(1 for t in daily_todos.get(abbr, []) if t.get('completed')) for abbr in day_abbrs)
            todo_pct = int((comp_todos / tot_todos * 100)) if tot_todos > 0 else 0

            tot_goals = len(goals)
            comp_goals = sum(1 for g in goals if g.get('completed'))

            tot_shopping = len(shopping_list)
            bought_shop = sum(1 for s in shopping_list if s.get('bought'))

            meals_p = sum(sum(1 for m_type in ['breakfast', 'lunch', 'dinner'] if meals_menu.get(abbr, {}).get(m_type)) for abbr in day_abbrs)
            meal_pct = int((meals_p / 21 * 100))

            s_comps = []
            if tot_todos > 0:
                s_comps.append(todo_pct)
            if tot_goals > 0:
                s_comps.append(int(comp_goals / tot_goals * 100))
            if meals_p > 0:
                s_comps.append(meal_pct)
            w_score = int(sum(s_comps) / len(s_comps)) if s_comps else (100 if comp_todos > 0 or comp_goals > 0 else 0)

            resp_data = {
                'success': True,
                'action': action,
                'completed_goals': comp_goals,
                'total_goals': tot_goals,
                'bought_shopping': bought_shop,
                'total_shopping': tot_shopping,
                'completed_todos': comp_todos,
                'total_todos': tot_todos,
                'todo_completion_pct': todo_pct,
                'weekly_score': w_score,
                'message': 'Operation completed successfully'
            }

            if action == 'add_weekly_goal' and goals:
                resp_data['goal'] = goals[-1]
                resp_data['message'] = 'Weekly goal added!'
            elif action == 'toggle_weekly_goal':
                goal_id = request.form.get('goal_id')
                g_comp = next((g.get('completed') for g in goals if g.get('id') == goal_id), False)
                resp_data['completed'] = g_comp
                resp_data['goal_id'] = goal_id
                resp_data['message'] = 'Weekly goal updated!'
            elif action == 'delete_weekly_goal':
                resp_data['goal_id'] = request.form.get('goal_id')
                resp_data['message'] = 'Weekly goal removed.'
            elif action == 'add_shopping_item' and shopping_list:
                resp_data['item'] = shopping_list[-1]
                resp_data['message'] = 'Shopping item added!'
            elif action == 'toggle_shopping_item':
                item_id = request.form.get('item_id')
                s_b = next((s.get('bought') for s in shopping_list if s.get('id') == item_id), False)
                resp_data['bought'] = s_b
                resp_data['item_id'] = item_id
                resp_data['message'] = 'Shopping item updated!'
            elif action == 'delete_shopping_item':
                resp_data['item_id'] = request.form.get('item_id')
                resp_data['message'] = 'Shopping item deleted.'
            elif action == 'add_daily_todo':
                day_abbr = request.form.get('day_abbr')
                if day_abbr in daily_todos and daily_todos[day_abbr]:
                    resp_data['todo'] = daily_todos[day_abbr][-1]
                resp_data['day_abbr'] = day_abbr
                resp_data['message'] = f'To-do added to {day_abbr}!'
            elif action == 'toggle_daily_todo':
                day_abbr = request.form.get('day_abbr')
                todo_id = request.form.get('todo_id')
                t_comp = False
                if day_abbr in daily_todos:
                    t_comp = next((t.get('completed') for t in daily_todos[day_abbr] if t.get('id') == todo_id), False)
                resp_data['completed'] = t_comp
                resp_data['todo_id'] = todo_id
                resp_data['day_abbr'] = day_abbr
                resp_data['message'] = 'To-do updated!'
            elif action == 'delete_daily_todo':
                resp_data['todo_id'] = request.form.get('todo_id')
                resp_data['day_abbr'] = request.form.get('day_abbr')
                resp_data['message'] = 'To-do item deleted.'

            return jsonify(resp_data)
        return redirect(url_for('planner.weekly', year=year_param, week=week_param))

    start_year = start_of_week.year
    end_year = end_of_week.year

    yearly_plans = YearlyPlan.query.filter(YearlyPlan.user_id == current_user.id, YearlyPlan.year.in_([start_year, end_year])).all()
    yearly_map = {yp.year: yp for yp in yearly_plans}

    month_keys = list({(start_of_week.year, start_of_week.month), (end_of_week.year, end_of_week.month)})
    monthly_plans = MonthlyPlan.query.filter(
        MonthlyPlan.user_id == current_user.id,
        db.or_(*[db.and_(MonthlyPlan.year == y, MonthlyPlan.month == m) for y, m in month_keys])
    ).all()
    monthly_map = {(mp.year, mp.month): mp for mp in monthly_plans}

    days_of_week = []
    for i in range(7):
        d = start_of_week + timedelta(days=i)
        yp = yearly_map.get(d.year)
        mp = monthly_map.get((d.year, d.month))
        days_of_week.append({
            'abbr': day_abbrs[i],
            'name': d.strftime('%A'),
            'date_str': d.strftime('%b %d'),
            'full_date': d.strftime('%Y-%m-%d'),
            'date_obj': d,
            'cascaded_items': get_all_cascaded_items_for_daily(current_user.id, d, yearly_plan=yp, monthly_plan=mp, weekly_plan=plan)
        })

    goals = plan.goals if plan and plan.goals else []
    daily_todos = plan.daily_todos if plan and plan.daily_todos else {abbr: [] for abbr in day_abbrs}
    shopping_list = plan.shopping_list if plan and plan.shopping_list else []
    meals_menu = plan.meals_menu if plan and plan.meals_menu else {abbr: {'breakfast': '', 'lunch': '', 'dinner': ''} for abbr in day_abbrs}
    notes = plan.notes if plan and plan.notes else ''

    for abbr in day_abbrs:
        if abbr not in daily_todos:
            daily_todos[abbr] = []
        if abbr not in meals_menu:
            meals_menu[abbr] = {'breakfast': '', 'lunch': '', 'dinner': ''}

    # Entire Week Report calculations
    total_todos = sum(len(daily_todos.get(abbr, [])) for abbr in day_abbrs)
    completed_todos = sum(sum(1 for t in daily_todos.get(abbr, []) if t.get('completed')) for abbr in day_abbrs)
    todo_completion_pct = int((completed_todos / total_todos * 100)) if total_todos > 0 else 0

    total_goals = len(goals)
    completed_goals = sum(1 for g in goals if g.get('completed'))

    total_shopping = len(shopping_list)
    bought_shopping = sum(1 for s in shopping_list if s.get('bought'))

    meals_planned = sum(sum(1 for m_type in ['breakfast', 'lunch', 'dinner'] if meals_menu.get(abbr, {}).get(m_type)) for abbr in day_abbrs)
    meal_planning_pct = int((meals_planned / 21 * 100))

    # Depression episode summary for this week
    week_episodes = []
    week_daily_plans = DailyPlan.query.filter(
        DailyPlan.user_id == current_user.id,
        DailyPlan.date >= start_of_week,
        DailyPlan.date <= end_of_week
    ).all()
    for dp in week_daily_plans:
        if dp.depression_episodes:
            for ep in dp.depression_episodes:
                ep_copy = dict(ep)
                ep_copy['date_str'] = dp.date.strftime('%b %d')
                week_episodes.append(ep_copy)

    score_components = []
    if total_todos > 0:
        score_components.append(todo_completion_pct)
    if total_goals > 0:
        score_components.append(int(completed_goals / total_goals * 100))
    if meals_planned > 0:
        score_components.append(meal_planning_pct)
    weekly_score = int(sum(score_components) / len(score_components)) if score_components else (100 if completed_todos > 0 or completed_goals > 0 else 0)

    # Daily Planner -> Weekly Summary Cascade
    daily_tasks_done = 0
    daily_tasks_total = 0
    daily_mood_tally = {}
    for dp in week_daily_plans:
        if dp.tasks:
            daily_tasks_total += len(dp.tasks)
            daily_tasks_done += sum(1 for t in dp.tasks if t.get('completed'))
        if dp.schedule:
            for slot, sdata in dp.schedule.items():
                if isinstance(sdata, dict) and sdata.get('mood'):
                    m = sdata.get('mood')
                    daily_mood_tally[m] = daily_mood_tally.get(m, 0) + 1

    prev_week_date = start_of_week - timedelta(days=7)
    prev_year, prev_week, _ = prev_week_date.isocalendar()
    next_week_date = start_of_week + timedelta(days=7)
    next_year, next_week, _ = next_week_date.isocalendar()

    return render_template(
        'planner/weekly.html',
        selected_year=year_param,
        selected_week=week_param,
        today=today,
        current_year=current_year,
        current_week=current_week,
        start_of_week=start_of_week,
        end_of_week=end_of_week,
        days_of_week=days_of_week,
        plan=plan,
        goals=goals,
        daily_todos=daily_todos,
        shopping_list=shopping_list,
        meals_menu=meals_menu,
        notes=notes,
        total_todos=total_todos,
        completed_todos=completed_todos,
        todo_completion_pct=todo_completion_pct,
        total_goals=total_goals,
        completed_goals=completed_goals,
        total_shopping=total_shopping,
        bought_shopping=bought_shopping,
        meals_planned=meals_planned,
        meal_planning_pct=meal_planning_pct,
        week_episodes=week_episodes,
        weekly_score=weekly_score,
        daily_tasks_done=daily_tasks_done,
        daily_tasks_total=daily_tasks_total,
        daily_mood_tally=daily_mood_tally,
        daily_planned_days=len(week_daily_plans),
        prev_year=prev_year,
        prev_week=prev_week,
        next_year=next_year,
        next_week=next_week
    )


@planner.route('/daily', methods=['GET', 'POST'])
@login_required
def daily():
    today = get_today_date()
    date_param = request.args.get('date')
    if date_param and date_param.lower() != 'today':
        try:
            selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
    else:
        selected_date = today

    # Automatically process spillover tasks and daily routine defaults up to selected_date
    process_task_spillovers(current_user.id, selected_date)
    populate_daily_defaults(current_user.id, selected_date)

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=selected_date).first()

    if request.method == 'POST':
        action = request.form.get('action')
        
        if not plan:
            plan = DailyPlan(user_id=current_user.id, date=selected_date, schedule={}, tasks=[], notes='')
            db.session.add(plan)

        tasks = plan.tasks or []
        schedule = plan.schedule or {}

        if action == 'add_task':
            task_text = request.form.get('task_text', '').strip()
            priority = request.form.get('priority', 'Medium')
            is_default = bool(request.form.get('is_default'))
            task_tags = request.form.getlist('tags') or (request.form.get('tags', '').split(',') if request.form.get('tags') else [])
            task_tags = [t.strip() for t in task_tags if t.strip()]

            if task_text:
                new_task = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'text': task_text,
                    'priority': priority,
                    'tags': task_tags,
                    'completed': False,
                    'is_default': is_default,
                    'is_spillover': False,
                    'spillover_count': 0,
                    'original_date': selected_date.strftime('%Y-%m-%d')
                }
                tasks.append(new_task)
                plan.tasks = tasks
                flag_modified(plan, 'tasks')

                # Clear previous deletion/dismissal records on this plan for this task text
                raw_sched = plan.schedule or {}
                if isinstance(raw_sched, dict):
                    t_key = task_text.lower()
                    deleted_defs = [d for d in raw_sched.get('_deleted_defaults', []) if str(d).lower() != t_key]
                    dismissed = [d for d in raw_sched.get('_dismissed_tasks', []) if str(d).lower() != t_key]
                    if len(deleted_defs) != len(raw_sched.get('_deleted_defaults', [])) or len(dismissed) != len(raw_sched.get('_dismissed_tasks', [])):
                        raw_sched['_deleted_defaults'] = deleted_defs
                        raw_sched['_dismissed_tasks'] = dismissed
                        plan.schedule = raw_sched
                        flag_modified(plan, 'schedule')

                db.session.commit()
                flash('Task added successfully!', 'success')

        elif action == 'delete_task':
            task_id = request.form.get('task_id')
            deleted_task = None
            for t in tasks:
                if isinstance(t, dict) and t.get('id') == task_id:
                    deleted_task = t
                    break

            plan.tasks = [t for t in tasks if isinstance(t, dict) and t.get('id') != task_id]
            
            raw_sched = plan.schedule or {}
            if not isinstance(raw_sched, dict):
                raw_sched = {}

            dismissed = list(raw_sched.get('_dismissed_tasks', []))
            deleted_defs = list(raw_sched.get('_deleted_defaults', []))

            if task_id not in dismissed:
                dismissed.append(task_id)

            if deleted_task:
                t_text = deleted_task.get('text', '').strip().lower()
                if t_text and t_text not in deleted_defs:
                    deleted_defs.append(t_text)
                if task_id not in deleted_defs:
                    deleted_defs.append(task_id)

            raw_sched['_dismissed_tasks'] = dismissed
            raw_sched['_deleted_defaults'] = deleted_defs
            plan.schedule = raw_sched
            flag_modified(plan, 'schedule')

            flag_modified(plan, 'tasks')
            db.session.commit()
            flash('Task deleted.', 'info')

        elif action == 'save_schedule':
            new_schedule = {}
            default_slots = [
                "05:00 - 06:00 AM", "06:00 - 07:00 AM", "07:00 - 08:00 AM", "08:00 - 09:00 AM",
                "09:00 - 10:00 AM", "10:00 - 11:00 AM", "11:00 - 12:00 PM", "12:00 - 01:00 PM",
                "01:00 - 02:00 PM", "02:00 - 03:00 PM", "03:00 - 04:00 PM", "04:00 - 05:00 PM",
                "05:00 - 06:00 PM", "06:00 - 07:00 PM", "07:00 - 08:00 PM", "08:00 - 09:00 PM",
                "09:00 - 10:00 PM", "10:00 - 11:00 PM", "11:00 - 12:00 AM"
            ]
            raw_sched = plan.schedule or {}
            if isinstance(raw_sched, dict):
                for k, v in raw_sched.items():
                    if k.startswith('_'):
                        new_schedule[k] = v

            for time_slot in default_slots:
                slot_id = time_slot.replace(':', '_').replace(' ', '_').replace('-', '_')
                meridiem = time_slot.split(' ')[-1] if ' ' in time_slot else ''
                start_part = time_slot.split(' - ')[0] if ' - ' in time_slot else time_slot
                start_time = f"{start_part} {meridiem}".strip() if meridiem and not start_part.endswith(meridiem) else start_part
                start_slot_id = start_time.replace(':', '_').replace(' ', '_')
                act = (request.form.get(f'slot_act_{slot_id}') or request.form.get(f'slot_act_{start_slot_id}') or request.form.get(f'slot_{time_slot}') or '').strip()
                mood = (request.form.get(f'slot_mood_{slot_id}') or request.form.get(f'slot_mood_{start_slot_id}') or '').strip()
                is_def = bool(request.form.get(f'slot_def_{slot_id}') or request.form.get(f'slot_def_{start_slot_id}'))

                if act or mood or is_def or (isinstance(raw_sched, dict) and time_slot in raw_sched):
                    new_schedule[time_slot] = {'activity': act, 'mood': mood, 'is_default': is_def}

            plan.schedule = new_schedule
            flag_modified(plan, 'schedule')
            db.session.commit()
            if not (request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('is_ajax') == 'true'):
                flash('Hourly activity & mood tracker saved successfully!', 'success')

        elif action == 'save_notes':
            notes = request.form.get('notes', '').strip()
            plan.notes = notes
            db.session.commit()
            if not (request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('is_ajax') == 'true'):
                flash('Notes updated!', 'success')

        elif action == 'add_depression_episode':
            start_time = request.form.get('start_time', '').strip() or 'N/A'
            duration = request.form.get('duration', '').strip() or 'N/A'
            try:
                intensity = int(request.form.get('intensity', 5))
            except ValueError:
                intensity = 5
            triggers = request.form.get('triggers', '').strip()
            coping_mechanism = request.form.get('coping_mechanism', '').strip()
            coping_effectiveness = request.form.get('coping_effectiveness', 'Helpful').strip()
            notes = request.form.get('notes', '').strip()
            entry_time = datetime.now().strftime('%I:%M %p')

            new_episode = {
                'id': str(int(datetime.utcnow().timestamp() * 1000)),
                'entry_time': entry_time,
                'start_time': start_time,
                'duration': duration,
                'intensity': intensity,
                'triggers': triggers,
                'coping_mechanism': coping_mechanism,
                'coping_effectiveness': coping_effectiveness,
                'notes': notes
            }

            episodes = plan.depression_episodes or []
            episodes.append(new_episode)
            plan.depression_episodes = episodes
            flag_modified(plan, 'depression_episodes')
            db.session.commit()
            flash('Depression episode & symptom logged.', 'success')

        elif action == 'delete_depression_episode':
            episode_id = request.form.get('episode_id')
            episodes = plan.depression_episodes or []
            plan.depression_episodes = [ep for ep in episodes if ep.get('id') != episode_id]
            flag_modified(plan, 'depression_episodes')
            db.session.commit()
            flash('Depression episode record deleted.', 'info')

        elif action == 'add_memory_log':
            time_val = request.form.get('time', '').strip() or 'N/A'
            item = request.form.get('item', '').strip()
            category = request.form.get('category', 'General').strip()
            context = request.form.get('context', '').strip()
            impact = request.form.get('impact', 'Mild').strip()
            recovery = request.form.get('recovery', 'Remembered later').strip()
            notes = request.form.get('notes', '').strip()
            entry_time = datetime.now().strftime('%I:%M %p')

            new_log = {
                'id': str(int(datetime.utcnow().timestamp() * 1000)),
                'entry_time': entry_time,
                'time': time_val,
                'item': item,
                'category': category,
                'context': context,
                'impact': impact,
                'recovery': recovery,
                'notes': notes
            }

            logs = plan.memory_logs or []
            logs.append(new_log)
            plan.memory_logs = logs
            flag_modified(plan, 'memory_logs')
            db.session.commit()
            flash('Memory slip logged successfully.', 'success')

        elif action == 'delete_memory_log':
            log_id = request.form.get('log_id')
            logs = plan.memory_logs or []
            plan.memory_logs = [m for m in logs if m.get('id') != log_id]
            flag_modified(plan, 'memory_logs')
            db.session.commit()
            flash('Memory log record deleted.', 'info')

        elif action == 'save_sleep_log':
            try:
                hours = float(request.form.get('sleep_hours', 7.0))
            except ValueError:
                hours = 7.0
            try:
                quality = int(request.form.get('sleep_quality', 8))
            except ValueError:
                quality = 8

            bedtime = request.form.get('bedtime', '').strip() or '11:00 PM'
            wake_time = request.form.get('wake_time', '').strip() or '07:00 AM'
            disruptions = request.form.get('disruptions', 'None').strip()
            notes = request.form.get('notes', '').strip()

            plan.sleep_log = {
                'hours': hours,
                'bedtime': bedtime,
                'wake_time': wake_time,
                'quality': quality,
                'disruptions': disruptions,
                'notes': notes,
                'updated_at': datetime.now().strftime('%I:%M %p')
            }
            flag_modified(plan, 'sleep_log')
            db.session.commit()
            flash('Sleep tracker metrics saved successfully.', 'success')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('is_ajax') == 'true':
            return jsonify({'success': True, 'action': action, 'message': 'Operation completed successfully'})
        return redirect(url_for('planner.daily', date=selected_date.strftime('%Y-%m-%d')))

    # Hourly default schedule slots (05:00 - 06:00 AM to 11:00 - 12:00 AM)
    default_slots = [
        "05:00 - 06:00 AM", "06:00 - 07:00 AM", "07:00 - 08:00 AM", "08:00 - 09:00 AM",
        "09:00 - 10:00 AM", "10:00 - 11:00 AM", "11:00 - 12:00 PM", "12:00 - 01:00 PM",
        "01:00 - 02:00 PM", "02:00 - 03:00 PM", "03:00 - 04:00 PM", "04:00 - 05:00 PM",
        "05:00 - 06:00 PM", "06:00 - 07:00 PM", "07:00 - 08:00 PM", "08:00 - 09:00 PM",
        "09:00 - 10:00 PM", "10:00 - 11:00 PM", "11:00 - 12:00 AM"
    ]

    # Normalize schedule dict for 12h format & mood tracking
    raw_schedule = plan.schedule if plan and plan.schedule else {}
    normalized_schedule = {}
    legacy_map = {f'{h:02d}:00': f'{(h if (h % 12 != 0) else 12):02d}:00 {"AM" if h < 12 else "PM"}' for h in range(5, 24)}

    for slot in default_slots:
        val = raw_schedule.get(slot)
        if not val:
            start_time = slot.split(' - ')[0] if ' - ' in slot else slot
            if start_time in raw_schedule:
                val = raw_schedule[start_time]
            else:
                for k24, v12 in legacy_map.items():
                    if v12 == start_time and k24 in raw_schedule:
                        val = raw_schedule[k24]
                        break

        if isinstance(val, dict):
            normalized_schedule[slot] = val
        elif isinstance(val, str) and val:
            normalized_schedule[slot] = {'activity': val, 'mood': ''}
        else:
            normalized_schedule[slot] = {'activity': '', 'mood': ''}
    
    cascaded_items = get_all_cascaded_items_for_daily(current_user.id, selected_date)
    user_tags = get_user_tags(current_user)

    raw_tasks = plan.tasks if plan and plan.tasks else []
    sorted_tasks = sorted(raw_tasks, key=lambda t: 1 if (isinstance(t, dict) and t.get('completed')) else 0)

    return render_template(
        'planner/daily.html',
        selected_date=selected_date,
        today=today,
        plan=plan,
        tasks=sorted_tasks,
        schedule=normalized_schedule,
        notes=plan.notes if plan else '',
        depression_episodes=plan.depression_episodes if plan and plan.depression_episodes else [],
        memory_logs=plan.memory_logs if plan and plan.memory_logs else [],
        sleep_log=plan.sleep_log if plan and plan.sleep_log else {},
        default_slots=default_slots,
        cascaded_items=cascaded_items,
        user_tags=user_tags
    )


@planner.route('/monthly', methods=['GET', 'POST'])
@login_required
def monthly():
    today = get_today_date()
    try:
        selected_year = int(request.args.get('year', today.year))
        selected_month = int(request.args.get('month', today.month))
    except (ValueError, TypeError):
        selected_year = today.year
        selected_month = today.month

    # Ensure valid month
    if selected_month < 1 or selected_month > 12:
        selected_month = today.month

    plan = MonthlyPlan.query.filter_by(user_id=current_user.id, year=selected_year, month=selected_month).first()

    if request.method == 'POST':
        action = request.form.get('action')
        
        if not plan:
            plan = MonthlyPlan(user_id=current_user.id, year=selected_year, month=selected_month, goals=[], habits=[], milestones=[], calendar_days={}, notes='')
            db.session.add(plan)

        goals = plan.goals or []
        habits = plan.habits or []
        milestones = plan.milestones or []
        calendar_days = plan.calendar_days or {}

        if action == 'add_goal':
            title = request.form.get('goal_title', '').strip()
            category = request.form.get('category', 'Personal')
            if title:
                new_goal = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'title': title,
                    'category': category,
                    'status': 'In Progress'
                }
                goals.append(new_goal)
                plan.goals = goals
                flag_modified(plan, 'goals')
                db.session.commit()
                flash('Goal added!', 'success')

        elif action == 'toggle_goal_status':
            goal_id = request.form.get('goal_id')
            for g in goals:
                if g.get('id') == goal_id:
                    g['status'] = 'Completed' if g.get('status') != 'Completed' else 'In Progress'
            plan.goals = goals
            flag_modified(plan, 'goals')
            db.session.commit()

        elif action == 'delete_goal':
            goal_id = request.form.get('goal_id')
            plan.goals = [g for g in goals if g.get('id') != goal_id]
            flag_modified(plan, 'goals')
            db.session.commit()
            flash('Goal removed.', 'info')

        elif action == 'add_habit':
            name = request.form.get('habit_name', '').strip()
            if name:
                new_habit = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'name': name,
                    'completed_days': []
                }
                habits.append(new_habit)
                plan.habits = habits
                flag_modified(plan, 'habits')
                db.session.commit()
                flash('Habit added!', 'success')

        elif action == 'delete_habit':
            habit_id = request.form.get('habit_id')
            plan.habits = [h for h in habits if h.get('id') != habit_id]
            flag_modified(plan, 'habits')
            db.session.commit()

        elif action == 'add_milestone':
            title = request.form.get('milestone_title', '').strip()
            target_day = request.form.get('target_day', '').strip()
            if title:
                new_ms = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'title': title,
                    'day': target_day,
                    'completed': False
                }
                milestones.append(new_ms)
                plan.milestones = milestones
                flag_modified(plan, 'milestones')
                db.session.commit()
                flash('Milestone added!', 'success')

        elif action == 'toggle_milestone':
            ms_id = request.form.get('milestone_id')
            for ms in milestones:
                if ms.get('id') == ms_id:
                    ms['completed'] = not ms.get('completed', False)
            plan.milestones = milestones
            flag_modified(plan, 'milestones')
            db.session.commit()

        elif action == 'delete_milestone':
            ms_id = request.form.get('milestone_id')
            plan.milestones = [m for m in milestones if m.get('id') != ms_id]
            flag_modified(plan, 'milestones')
            db.session.commit()

        elif action == 'add_calendar_item':
            day_str = str(request.form.get('day', '')).strip()
            item_text = request.form.get('item_text', '').strip()
            item_type = request.form.get('item_type', 'deadline')
            sticker = request.form.get('sticker', '').strip()
            image_url = request.form.get('image_url', '').strip()
            remind_me = request.form.get('remind_me') in ['true', 'True', '1', 'on', True]

            if day_str and item_text:
                if day_str not in calendar_days:
                    calendar_days[day_str] = {'items': [], 'sticker': '', 'image_url': ''}
                
                day_entry = calendar_days[day_str]
                items = day_entry.get('items', [])
                items.append({
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'text': item_text,
                    'type': item_type,
                    'sticker': sticker,
                    'image_url': image_url,
                    'remind_me': remind_me
                })
                day_entry['items'] = items
                if sticker:
                    day_entry['sticker'] = sticker
                if image_url:
                    day_entry['image_url'] = image_url

                calendar_days[day_str] = day_entry
                plan.calendar_days = calendar_days
                flag_modified(plan, 'calendar_days')
                db.session.commit()
                flash(f'Plan item added to Day {day_str}!', 'success')

        elif action == 'edit_calendar_item':
            day_str = str(request.form.get('day', '')).strip()
            item_id = request.form.get('item_id', '').strip()
            item_text = request.form.get('item_text', '').strip()
            item_type = request.form.get('item_type', 'deadline')
            sticker = request.form.get('sticker', '').strip()
            image_url = request.form.get('image_url', '').strip()
            remind_me = request.form.get('remind_me') in ['true', 'True', '1', 'on', True]

            if day_str in calendar_days and item_id and item_text:
                day_entry = calendar_days[day_str]
                items = day_entry.get('items', [])
                for item in items:
                    if item.get('id') == item_id:
                        item['text'] = item_text
                        item['type'] = item_type
                        item['sticker'] = sticker
                        item['image_url'] = image_url
                        item['remind_me'] = remind_me
                        break
                day_entry['items'] = items

                remaining_stickers = [i.get('sticker') for i in items if i.get('sticker')]
                remaining_images = [i.get('image_url') for i in items if i.get('image_url')]
                day_entry['sticker'] = remaining_stickers[-1] if remaining_stickers else ''
                day_entry['image_url'] = remaining_images[-1] if remaining_images else ''

                calendar_days[day_str] = day_entry
                plan.calendar_days = calendar_days
                flag_modified(plan, 'calendar_days')
                db.session.commit()
                flash(f'Plan item updated on Day {day_str}!', 'success')

        elif action == 'delete_calendar_item':
            day_str = str(request.form.get('day', '')).strip()
            item_id = request.form.get('item_id')
            if day_str in calendar_days:
                day_entry = calendar_days[day_str]
                day_entry['items'] = [i for i in day_entry.get('items', []) if i.get('id') != item_id]
                
                remaining_stickers = [i.get('sticker') for i in day_entry['items'] if i.get('sticker')]
                remaining_images = [i.get('image_url') for i in day_entry['items'] if i.get('image_url')]
                day_entry['sticker'] = remaining_stickers[-1] if remaining_stickers else ''
                day_entry['image_url'] = remaining_images[-1] if remaining_images else ''

                calendar_days[day_str] = day_entry
                plan.calendar_days = calendar_days
                flag_modified(plan, 'calendar_days')
                db.session.commit()
                flash('Calendar item removed.', 'info')

        elif action == 'delete_day_sticker':
            day_str = str(request.form.get('day', '')).strip()
            if day_str in calendar_days:
                calendar_days[day_str]['sticker'] = ''
                calendar_days[day_str]['image_url'] = ''
                plan.calendar_days = calendar_days
                flag_modified(plan, 'calendar_days')
                db.session.commit()
                flash(f'Day sticker cleared for Day {day_str}.', 'info')

        elif action == 'set_day_sticker':
            day_str = str(request.form.get('day', '')).strip()
            sticker = request.form.get('sticker', '').strip()
            image_url = request.form.get('image_url', '').strip()
            if day_str:
                if day_str not in calendar_days:
                    calendar_days[day_str] = {'items': [], 'sticker': '', 'image_url': ''}
                calendar_days[day_str]['sticker'] = sticker
                if image_url:
                    calendar_days[day_str]['image_url'] = image_url
                plan.calendar_days = calendar_days
                flag_modified(plan, 'calendar_days')
                db.session.commit()
                flash(f'Sticker updated for Day {day_str}!', 'success')

        elif action == 'save_notes':
            notes = request.form.get('notes', '').strip()
            plan.notes = notes
            db.session.commit()
            flash('Monthly notes updated!', 'success')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('is_ajax') == 'true':
            tot_goals = len(goals)
            comp_goals = sum(1 for g in goals if g.get('status') == 'Completed')

            resp_data = {
                'success': True,
                'action': action,
                'completed_goals': comp_goals,
                'total_goals': tot_goals,
                'message': 'Operation completed successfully'
            }

            if action == 'add_goal' and goals:
                resp_data['goal'] = goals[-1]
                resp_data['message'] = 'Goal added!'
            elif action == 'toggle_goal_status':
                goal_id = request.form.get('goal_id')
                g_status = next((g.get('status') for g in goals if g.get('id') == goal_id), 'In Progress')
                resp_data['goal_id'] = goal_id
                resp_data['status'] = g_status
                resp_data['message'] = 'Goal status updated!'
            elif action == 'delete_goal':
                resp_data['goal_id'] = request.form.get('goal_id')
                resp_data['message'] = 'Goal removed.'
            elif action == 'add_milestone' and milestones:
                resp_data['milestone'] = milestones[-1]
                resp_data['message'] = 'Milestone added!'
            elif action == 'toggle_milestone':
                ms_id = request.form.get('milestone_id')
                ms_comp = next((m.get('completed') for m in milestones if m.get('id') == ms_id), False)
                resp_data['milestone_id'] = ms_id
                resp_data['completed'] = ms_comp
                resp_data['message'] = 'Milestone updated!'
            elif action == 'delete_milestone':
                resp_data['milestone_id'] = request.form.get('milestone_id')
                resp_data['message'] = 'Milestone deleted.'
            elif action in ['add_calendar_item', 'edit_calendar_item']:
                day_str = str(request.form.get('day', '')).strip()
                resp_data['day'] = day_str
                resp_data['day_entry'] = calendar_days.get(day_str, {'items': [], 'sticker': '', 'image_url': ''})
                resp_data['message'] = 'Calendar plan saved!'
            elif action == 'delete_calendar_item':
                day_str = str(request.form.get('day', '')).strip()
                resp_data['day'] = day_str
                resp_data['item_id'] = request.form.get('item_id')
                resp_data['day_entry'] = calendar_days.get(day_str, {'items': [], 'sticker': '', 'image_url': ''})
                resp_data['message'] = 'Calendar item removed.'
            elif action == 'delete_day_sticker':
                day_str = str(request.form.get('day', '')).strip()
                resp_data['day'] = day_str
                resp_data['message'] = f'Day sticker cleared for Day {day_str}.'
            elif action == 'add_habit' and habits:
                days_in_m = calendar.monthrange(selected_year, selected_month)[1]
                resp_data['habit'] = habits[-1]
                resp_data['days_in_month'] = days_in_m
                resp_data['message'] = 'Habit added!'
            elif action == 'delete_habit':
                resp_data['habit_id'] = request.form.get('habit_id')
                resp_data['message'] = 'Habit deleted.'

            return jsonify(resp_data)
        return redirect(url_for('planner.monthly', year=selected_year, month=selected_month))

    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    month_name = calendar.month_name[selected_month]

    # Calculate exact calendar weeks for grid view (Monday to Sunday)
    cal = calendar.Calendar(firstweekday=0)
    month_weeks = cal.monthdatescalendar(selected_year, selected_month)

    # Weekly Summaries Cascade -> Monthly Planner
    first_day_of_month = date(selected_year, selected_month, 1)
    last_day_of_month = date(selected_year, selected_month, days_in_month)
    month_weekly_plans = WeeklyPlan.query.filter(
        WeeklyPlan.user_id == current_user.id,
        WeeklyPlan.start_date >= (first_day_of_month - timedelta(days=6)),
        WeeklyPlan.start_date <= last_day_of_month
    ).order_by(WeeklyPlan.start_date.asc()).all()

    weekly_summaries_list = []
    for wp in month_weekly_plans:
        w_goals = wp.goals or []
        w_goals_completed = sum(1 for g in w_goals if g.get('completed'))
        w_todos = wp.daily_todos or {}
        w_todos_total = sum(len(w_todos.get(d, [])) for d in ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'])
        w_todos_completed = sum(sum(1 for t in w_todos.get(d, []) if t.get('completed')) for d in ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'])

        weekly_summaries_list.append({
            'week_number': wp.week_number,
            'start_date_str': wp.start_date.strftime('%b %d'),
            'goals_done': f"{w_goals_completed}/{len(w_goals)}",
            'todos_done': f"{w_todos_completed}/{w_todos_total}",
            'notes': wp.notes or ''
        })

    cascaded_yearly_events = get_yearly_events_for_month(current_user.id, selected_year, selected_month)

    return render_template(
        'planner/monthly.html',
        selected_year=selected_year,
        selected_month=selected_month,
        month_name=month_name,
        days_in_month=days_in_month,
        month_weeks=month_weeks,
        today=today,
        plan=plan,
        goals=plan.goals if plan else [],
        habits=plan.habits if plan else [],
        milestones=plan.milestones if plan else [],
        calendar_days=plan.calendar_days if (plan and plan.calendar_days) else {},
        notes=plan.notes if plan else '',
        weekly_summaries_list=weekly_summaries_list,
        cascaded_yearly_events=cascaded_yearly_events,
        calendar=calendar
    )


@planner.route('/yearly', methods=['GET', 'POST'])
@login_required
def yearly():
    today = get_today_date()
    try:
        selected_year = int(request.args.get('year', today.year))
    except (ValueError, TypeError):
        selected_year = today.year

    plan = YearlyPlan.query.filter_by(user_id=current_user.id, year=selected_year).first()

    if request.method == 'POST':
        action = request.form.get('action')
        
        if not plan:
            plan = YearlyPlan(user_id=current_user.id, year=selected_year, resolutions=[], objectives=[], events=[], reflections='')
            db.session.add(plan)

        resolutions = plan.resolutions or []
        objectives = plan.objectives or []
        events = plan.events or []

        if action == 'add_resolution':
            text = request.form.get('resolution_text', '').strip()
            category = request.form.get('category', 'Personal')
            if text:
                new_res = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'text': text,
                    'category': category,
                    'completed': False
                }
                resolutions.append(new_res)
                plan.resolutions = resolutions
                flag_modified(plan, 'resolutions')
                db.session.commit()
                flash('Resolution added!', 'success')

        elif action == 'toggle_resolution':
            res_id = request.form.get('resolution_id')
            for r in resolutions:
                if r.get('id') == res_id:
                    r['completed'] = not r.get('completed', False)
            plan.resolutions = resolutions
            flag_modified(plan, 'resolutions')
            db.session.commit()

        elif action == 'delete_resolution':
            res_id = request.form.get('resolution_id')
            plan.resolutions = [r for r in resolutions if r.get('id') != res_id]
            flag_modified(plan, 'resolutions')
            db.session.commit()

        elif action == 'add_objective':
            title = request.form.get('objective_title', '').strip()
            quarter = request.form.get('quarter', 'Q1')
            if title:
                new_obj = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'title': title,
                    'quarter': quarter,
                    'status': 'In Progress'
                }
                objectives.append(new_obj)
                plan.objectives = objectives
                flag_modified(plan, 'objectives')
                db.session.commit()
                flash('Objective added!', 'success')

        elif action == 'update_objective_status':
            obj_id = request.form.get('objective_id')
            new_status = request.form.get('status', 'In Progress')
            for o in objectives:
                if o.get('id') == obj_id:
                    o['status'] = new_status
            plan.objectives = objectives
            flag_modified(plan, 'objectives')
            db.session.commit()

        elif action == 'delete_objective':
            obj_id = request.form.get('objective_id')
            plan.objectives = [o for o in objectives if o.get('id') != obj_id]
            flag_modified(plan, 'objectives')
            db.session.commit()

        elif action == 'add_yearly_event':
            title = request.form.get('event_title', '').strip()
            event_type = request.form.get('event_type', 'goal')
            event_date = request.form.get('event_date', '').strip()
            notes = request.form.get('notes', '').strip()
            if title and event_date:
                new_event = {
                    'id': str(int(datetime.utcnow().timestamp() * 1000)),
                    'title': title,
                    'event_type': event_type,
                    'date': event_date,
                    'notes': notes,
                    'completed': False
                }
                events.append(new_event)
                plan.events = events
                flag_modified(plan, 'events')
                db.session.commit()
                flash('Yearly event/goal added!', 'success')

        elif action == 'toggle_yearly_event':
            event_id = request.form.get('event_id')
            for ev in events:
                if ev.get('id') == event_id:
                    ev['completed'] = not ev.get('completed', False)
            plan.events = events
            flag_modified(plan, 'events')
            db.session.commit()

        elif action == 'delete_yearly_event':
            event_id = request.form.get('event_id')
            plan.events = [ev for ev in events if ev.get('id') != event_id]
            flag_modified(plan, 'events')
            db.session.commit()
            flash('Yearly event removed.', 'info')

        elif action == 'save_reflections':
            reflections = request.form.get('reflections', '').strip()
            plan.reflections = reflections
            db.session.commit()
            flash('Yearly reflections saved!', 'success')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('is_ajax') == 'true':
            tot_events = len(events)
            tot_res = len(resolutions)
            acc_res = sum(1 for r in resolutions if r.get('completed'))

            resp_data = {
                'success': True,
                'action': action,
                'total_events': tot_events,
                'total_resolutions': tot_res,
                'accomplished_resolutions': acc_res,
                'message': 'Operation completed successfully'
            }

            if action == 'add_yearly_event' and events:
                resp_data['event'] = events[-1]
                resp_data['message'] = 'Yearly event/goal added!'
            elif action == 'toggle_yearly_event':
                event_id = request.form.get('event_id')
                ev_comp = next((ev.get('completed') for ev in events if ev.get('id') == event_id), False)
                resp_data['event_id'] = event_id
                resp_data['completed'] = ev_comp
                resp_data['message'] = 'Yearly event updated!'
            elif action == 'delete_yearly_event':
                resp_data['event_id'] = request.form.get('event_id')
                resp_data['message'] = 'Yearly event removed.'
            elif action == 'add_resolution' and resolutions:
                resp_data['resolution'] = resolutions[-1]
                resp_data['message'] = 'Resolution added!'
            elif action == 'toggle_resolution':
                res_id = request.form.get('resolution_id')
                res_comp = next((r.get('completed') for r in resolutions if r.get('id') == res_id), False)
                resp_data['resolution_id'] = res_id
                resp_data['completed'] = res_comp
                resp_data['message'] = 'Resolution updated!'
            elif action == 'delete_resolution':
                resp_data['resolution_id'] = request.form.get('resolution_id')
                resp_data['message'] = 'Resolution removed.'
            elif action == 'add_objective' and objectives:
                resp_data['objective'] = objectives[-1]
                resp_data['message'] = 'Objective added!'
            elif action == 'update_objective_status':
                obj_id = request.form.get('objective_id')
                obj_stat = request.form.get('status', 'In Progress')
                resp_data['objective_id'] = obj_id
                resp_data['status'] = obj_stat
                resp_data['message'] = 'Objective status updated!'
            elif action == 'delete_objective':
                resp_data['objective_id'] = request.form.get('objective_id')
                resp_data['message'] = 'Objective deleted.'
            elif action == 'save_reflections':
                resp_data['message'] = 'Annual reflections saved!'

            return jsonify(resp_data)
        return redirect(url_for('planner.yearly', year=selected_year))

    # 12-Month Achievement Grid Summary Cascade (Monthly -> Yearly)
    all_monthly_plans = MonthlyPlan.query.filter_by(user_id=current_user.id, year=selected_year).all()
    monthly_plan_by_month = {mp.month: mp for mp in all_monthly_plans}

    months_achievement_grid = []
    for m in range(1, 13):
        m_name = calendar.month_name[m]
        mp = monthly_plan_by_month.get(m)

        m_goals = mp.goals if mp and mp.goals else []
        m_goals_completed = sum(1 for g in m_goals if g.get('status') == 'Completed')
        m_goals_total = len(m_goals)

        m_milestones = mp.milestones if mp and mp.milestones else []
        m_milestones_completed = sum(1 for ms in m_milestones if ms.get('completed'))
        m_milestones_total = len(m_milestones)

        m_habits = mp.habits if mp and mp.habits else []
        m_habits_count = len(m_habits)

        components = []
        if m_goals_total > 0:
            components.append(int(m_goals_completed / m_goals_total * 100))
        if m_milestones_total > 0:
            components.append(int(m_milestones_completed / m_milestones_total * 100))

        achievement_score = int(sum(components) / len(components)) if components else (100 if m_goals_completed > 0 or m_milestones_completed > 0 else 0)

        months_achievement_grid.append({
            'month_num': m,
            'month_name': m_name,
            'has_plan': mp is not None,
            'goals_done': m_goals_completed,
            'goals_total': m_goals_total,
            'milestones_done': m_milestones_completed,
            'milestones_total': m_milestones_total,
            'habits_count': m_habits_count,
            'achievement_score': achievement_score
        })

    return render_template(
        'planner/yearly.html',
        selected_year=selected_year,
        today=today,
        plan=plan,
        resolutions=plan.resolutions if plan else [],
        objectives=plan.objectives if plan else [],
        events=plan.events if (plan and plan.events) else [],
        reflections=plan.reflections if plan else '',
        months_achievement_grid=months_achievement_grid
    )


# AJAX Endpoint for dynamic toggling of tasks
@planner.route('/api/daily/task/toggle', methods=['POST'])
@login_required
def api_toggle_task():
    data = request.get_json() or {}
    date_str = data.get('date')
    task_id = data.get('task_id')

    if not date_str or not task_id:
        return jsonify({'success': False, 'message': 'Missing arguments'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        return jsonify({'success': False, 'message': 'Plan not found'}), 404

    tasks = plan.tasks or []
    updated = False
    new_state = False

    for t in tasks:
        if t.get('id') == task_id:
            t['completed'] = not t.get('completed', False)
            new_state = t['completed']
            updated = True
            break

    if updated:
        plan.tasks = tasks
        flag_modified(plan, 'tasks')
        db.session.commit()
        return jsonify({'success': True, 'completed': new_state})
    
    return jsonify({'success': False, 'message': 'Task not found'}), 404


# AJAX Endpoint for dynamic reordering of tasks
@planner.route('/api/daily/task/reorder', methods=['POST'])
@login_required
def api_reorder_tasks():
    data = request.get_json() or {}
    date_str = data.get('date')
    task_ids = data.get('task_ids', [])

    if not date_str or not isinstance(task_ids, list):
        return jsonify({'success': False, 'message': 'Missing or invalid arguments'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        return jsonify({'success': False, 'message': 'Plan not found'}), 404

    tasks = plan.tasks or []
    task_map = {t.get('id'): t for t in tasks if isinstance(t, dict) and t.get('id')}

    reordered = []
    for tid in task_ids:
        if tid in task_map:
            reordered.append(task_map.pop(tid))

    # Append any remaining tasks that weren't specified
    reordered.extend(task_map.values())

    plan.tasks = reordered
    flag_modified(plan, 'tasks')
    db.session.commit()
    return jsonify({'success': True})


# AJAX Endpoint for dynamic task editing
@planner.route('/api/daily/task/edit', methods=['POST'])
@login_required
def api_edit_task():
    data = request.get_json() or {}
    date_str = data.get('date')
    task_id = data.get('task_id')
    new_text = data.get('text', '').strip() if 'text' in data else None
    new_priority = data.get('priority')
    is_default = bool(data.get('is_default')) if 'is_default' in data else None
    new_tags = data.get('tags')
    new_status = data.get('status')
    new_note = data.get('note')

    if not date_str or not task_id:
        return jsonify({'success': False, 'message': 'Missing required fields'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        return jsonify({'success': False, 'message': 'Plan not found'}), 404

    tasks = plan.tasks or []
    updated = False

    for t in tasks:
        if isinstance(t, dict) and t.get('id') == task_id:
            old_text = t.get('text', '').strip().lower()
            old_is_default = t.get('is_default', False)

            if new_text is not None:
                t['text'] = new_text
            if new_priority is not None:
                t['priority'] = new_priority
            if is_default is not None:
                t['is_default'] = is_default
                if old_is_default and not is_default:
                    raw_sched = plan.schedule or {}
                    if not isinstance(raw_sched, dict):
                        raw_sched = {}
                    deleted_defs = list(raw_sched.get('_deleted_defaults', []))
                    if old_text and old_text not in deleted_defs:
                        deleted_defs.append(old_text)
                    if task_id not in deleted_defs:
                        deleted_defs.append(task_id)
                    raw_sched['_deleted_defaults'] = deleted_defs
                    plan.schedule = raw_sched
                    flag_modified(plan, 'schedule')
                elif is_default:
                    raw_sched = plan.schedule or {}
                    if isinstance(raw_sched, dict):
                        t_key = (new_text or t.get('text', '')).lower()
                        deleted_defs = [d for d in raw_sched.get('_deleted_defaults', []) if str(d).lower() != t_key and str(d) != task_id]
                        dismissed = [d for d in raw_sched.get('_dismissed_tasks', []) if str(d).lower() != t_key and str(d) != task_id]
                        raw_sched['_deleted_defaults'] = deleted_defs
                        raw_sched['_dismissed_tasks'] = dismissed
                        plan.schedule = raw_sched
                        flag_modified(plan, 'schedule')
            if new_tags is not None:
                if isinstance(new_tags, str):
                    new_tags = [x.strip() for x in new_tags.split(',') if x.strip()]
                t['tags'] = new_tags
            if new_status is not None:
                st = str(new_status).strip()
                t['status'] = st
                if st == 'Completed':
                    t['completed'] = True
                elif st in ['To Do', 'In Progress', 'Undone']:
                    t['completed'] = False
            if new_note is not None:
                t['note'] = str(new_note).strip()
            updated = True
            break

    if updated:
        plan.tasks = tasks
        flag_modified(plan, 'tasks')
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Task not found'}), 404


# AJAX Endpoint for dynamic task addition
@planner.route('/api/daily/task/add', methods=['POST'])
@login_required
def api_add_task():
    data = request.get_json() or {}
    date_str = data.get('date')
    task_text = data.get('text', '').strip()
    priority = data.get('priority', 'Medium')
    is_default = bool(data.get('is_default'))
    tags = data.get('tags', [])
    note = data.get('note', '').strip()
    status = data.get('status', 'To Do')
    if isinstance(tags, str):
        tags = [x.strip() for x in tags.split(',') if x.strip()]

    if not date_str or not task_text:
        return jsonify({'success': False, 'message': 'Task text is required'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        plan = DailyPlan(user_id=current_user.id, date=target_date, schedule={}, tasks=[], notes='')
        db.session.add(plan)

    tasks = list(plan.tasks or [])
    new_task = {
        'id': str(int(datetime.utcnow().timestamp() * 1000)),
        'text': task_text,
        'priority': priority,
        'tags': tags,
        'status': status,
        'note': note,
        'completed': False,
        'is_default': is_default,
        'is_spillover': False,
        'spillover_count': 0,
        'original_date': target_date.strftime('%Y-%m-%d')
    }
    tasks.append(new_task)
    plan.tasks = tasks
    flag_modified(plan, 'tasks')

    # Clear previous deletion/dismissal records on this plan for this task text
    raw_sched = plan.schedule or {}
    if isinstance(raw_sched, dict):
        t_key = task_text.lower()
        deleted_defs = [d for d in raw_sched.get('_deleted_defaults', []) if str(d).lower() != t_key]
        dismissed = [d for d in raw_sched.get('_dismissed_tasks', []) if str(d).lower() != t_key]
        if len(deleted_defs) != len(raw_sched.get('_deleted_defaults', [])) or len(dismissed) != len(raw_sched.get('_dismissed_tasks', [])):
            raw_sched['_deleted_defaults'] = deleted_defs
            raw_sched['_dismissed_tasks'] = dismissed
            plan.schedule = raw_sched
            flag_modified(plan, 'schedule')

    db.session.commit()
    return jsonify({'success': True, 'task': new_task})


# AJAX Endpoint for dynamic task duplication
@planner.route('/api/daily/task/duplicate', methods=['POST'])
@login_required
def api_duplicate_task():
    data = request.get_json() or {}
    date_str = data.get('date')
    task_id = data.get('task_id')

    if not date_str or not task_id:
        return jsonify({'success': False, 'message': 'Missing required fields'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        return jsonify({'success': False, 'message': 'Plan not found'}), 404

    tasks = plan.tasks or []
    source_task = None
    for t in tasks:
        if isinstance(t, dict) and t.get('id') == task_id:
            source_task = t
            break

    if not source_task:
        return jsonify({'success': False, 'message': 'Source task not found'}), 404

    new_task = dict(source_task)
    new_task['id'] = str(int(datetime.utcnow().timestamp() * 1000))
    new_task['completed'] = False
    new_task['status'] = 'To Do'
    
    tasks.append(new_task)
    plan.tasks = tasks
    flag_modified(plan, 'tasks')
    db.session.commit()

    return jsonify({'success': True, 'task': new_task})


# AJAX Endpoint for dynamic task deletion
@planner.route('/api/daily/task/delete', methods=['POST'])
@login_required
def api_delete_task():
    data = request.get_json() or {}
    date_str = data.get('date')
    task_id = data.get('task_id')

    if not date_str or not task_id:
        return jsonify({'success': False, 'message': 'Missing task_id or date'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        return jsonify({'success': False, 'message': 'Plan not found'}), 404

    tasks = plan.tasks or []
    deleted_task = None
    for t in tasks:
        if isinstance(t, dict) and t.get('id') == task_id:
            deleted_task = t
            break

    if not deleted_task:
        return jsonify({'success': False, 'message': 'Task not found'}), 404

    plan.tasks = [t for t in tasks if isinstance(t, dict) and t.get('id') != task_id]
    
    raw_sched = plan.schedule or {}
    if not isinstance(raw_sched, dict):
        raw_sched = {}

    dismissed = list(raw_sched.get('_dismissed_tasks', []))
    deleted_defs = list(raw_sched.get('_deleted_defaults', []))

    if task_id not in dismissed:
        dismissed.append(task_id)

    t_text = deleted_task.get('text', '').strip().lower()
    if t_text and t_text not in deleted_defs:
        deleted_defs.append(t_text)
    if task_id not in deleted_defs:
        deleted_defs.append(task_id)

    raw_sched['_dismissed_tasks'] = dismissed
    raw_sched['_deleted_defaults'] = deleted_defs
    plan.schedule = raw_sched
    flag_modified(plan, 'schedule')
    flag_modified(plan, 'tasks')
    db.session.commit()

    return jsonify({'success': True})


# AJAX Endpoint for Custom Tag Management
@planner.route('/api/tags', methods=['GET', 'POST'])
@login_required
def api_user_tags():
    if request.method == 'GET':
        return jsonify({'success': True, 'tags': get_user_tags(current_user)})

    data = request.get_json() or {}
    action = data.get('action', 'add')
    user_tags = list(get_user_tags(current_user))

    if action == 'add':
        name = data.get('name', '').strip()
        color = data.get('color', '#3b82f6').strip()
        if name:
            new_tag = {
                'id': f'tag_{int(datetime.utcnow().timestamp() * 1000)}',
                'name': name,
                'color': color
            }
            user_tags.append(new_tag)
            current_user.custom_tags = user_tags
            flag_modified(current_user, 'custom_tags')
            db.session.commit()
            return jsonify({'success': True, 'tag': new_tag, 'tags': user_tags})
        return jsonify({'success': False, 'message': 'Tag name is required'}), 400

    elif action == 'delete':
        tag_id = data.get('tag_id')
        user_tags = [t for t in user_tags if t.get('id') != tag_id]
        current_user.custom_tags = user_tags
        flag_modified(current_user, 'custom_tags')
        db.session.commit()
        return jsonify({'success': True, 'tags': user_tags})

    elif action == 'edit':
        tag_id = data.get('tag_id')
        name = data.get('name', '').strip()
        color = data.get('color', '#3b82f6').strip()
        for t in user_tags:
            if t.get('id') == tag_id:
                if name:
                    t['name'] = name
                if color:
                    t['color'] = color
                break
        current_user.custom_tags = user_tags
        flag_modified(current_user, 'custom_tags')
        db.session.commit()
        return jsonify({'success': True, 'tags': user_tags})

    return jsonify({'success': False, 'message': 'Invalid tag action'}), 400


# AJAX Endpoint for real-time background schedule & mood slot updates
@planner.route('/api/daily/schedule/update', methods=['POST'])
@login_required
def api_update_schedule():
    data = request.get_json() or {}
    date_str = data.get('date')
    slot = data.get('slot')
    activity = data.get('activity', '').strip()
    mood = data.get('mood', '').strip()
    is_default = bool(data.get('is_default'))

    if not date_str or not slot:
        return jsonify({'success': False, 'message': 'Missing date or slot'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        plan = DailyPlan(user_id=current_user.id, date=target_date, schedule={}, tasks=[], notes='')
        db.session.add(plan)

    schedule = dict(plan.schedule or {})
    schedule[slot] = {
        'activity': activity,
        'mood': mood,
        'is_default': is_default
    }

    plan.schedule = schedule
    flag_modified(plan, 'schedule')
    db.session.commit()

    return jsonify({'success': True})


# AJAX Endpoint for real-time background daily notes updates
@planner.route('/api/daily/notes/update', methods=['POST'])
@login_required
def api_update_notes():
    data = request.get_json() or {}
    date_str = data.get('date')
    notes = data.get('notes', '').strip()

    if not date_str:
        return jsonify({'success': False, 'message': 'Missing date'}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid date format'}), 400

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()
    if not plan:
        plan = DailyPlan(user_id=current_user.id, date=target_date, schedule={}, tasks=[], notes='')
        db.session.add(plan)

    plan.notes = notes
    db.session.commit()

    return jsonify({'success': True})


# AJAX Endpoint for habit tracker day toggle
@planner.route('/api/monthly/habit/toggle', methods=['POST'])
@login_required
def api_toggle_habit_day():
    data = request.get_json() or {}
    year = data.get('year')
    month = data.get('month')
    habit_id = data.get('habit_id')
    day = data.get('day')

    if not all([year, month, habit_id, day]):
        return jsonify({'success': False, 'message': 'Missing arguments'}), 400

    plan = MonthlyPlan.query.filter_by(user_id=current_user.id, year=int(year), month=int(month)).first()
    if not plan:
        return jsonify({'success': False, 'message': 'Plan not found'}), 404

    habits = plan.habits or []
    updated = False
    is_checked = False

    for h in habits:
        if h.get('id') == habit_id:
            completed = h.get('completed_days', [])
            if day in completed:
                completed.remove(day)
                is_checked = False
            else:
                completed.append(day)
                is_checked = True
            h['completed_days'] = completed
            updated = True
            break

    if updated:
        plan.habits = habits
        flag_modified(plan, 'habits')
        db.session.commit()
        return jsonify({'success': True, 'checked': is_checked})

    return jsonify({'success': False, 'message': 'Habit not found'}), 404


# Google Drive Backup Sync Endpoint
@planner.route('/api/google/drive/sync', methods=['POST'])
@login_required
def api_google_drive_sync():
    from app.services.google_service import sync_to_google_drive
    try:
        res = sync_to_google_drive(current_user)
        last_sync = current_user.last_drive_sync.strftime('%Y-%m-%d %H:%M:%S UTC') if current_user.last_drive_sync else 'Just now'
        return jsonify({
            'success': res.get('success', False),
            'message': res.get('message', 'Drive sync completed'),
            'last_sync': last_sync,
            'google_connected': res.get('google_connected', True)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Drive sync error: {str(e)}'}), 500


# Google Drive Data Restore Endpoint
@planner.route('/api/google/drive/restore', methods=['POST'])
@login_required
def api_google_drive_restore():
    from app.services.google_service import restore_from_google_drive
    try:
        res = restore_from_google_drive(current_user)
        if res.get('success'):
            flash('Successfully restored planner data from Google Drive!', 'success')
            return jsonify({'success': True, 'message': res.get('message')})
        return jsonify({'success': False, 'message': res.get('message', 'Drive restore failed')}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': f'Drive restore error: {str(e)}'}), 500


# Direct Local JSON Backup Export Endpoint
@planner.route('/api/backup/export_json')
@login_required
def api_backup_export_json():
    from app.services.google_service import export_user_data_payload
    payload = export_user_data_payload(current_user)
    json_bytes = io.BytesIO(json.dumps(payload, indent=2).encode('utf-8'))
    filename = f"Chronos_Planner_Backup_{current_user.username}.json"
    return send_file(json_bytes, download_name=filename, as_attachment=True, mimetype="application/json")


# Direct Local JSON Backup Restore Endpoint
@planner.route('/api/backup/restore_json', methods=['POST'])
@login_required
def api_backup_restore_json():
    from app.services.google_service import import_user_data_payload
    try:
        if 'backup_file' in request.files:
            file = request.files['backup_file']
            if not file.filename or file.filename == '':
                return jsonify({'success': False, 'message': 'No backup file selected'}), 400
            content = file.read().decode('utf-8')
            payload = json.loads(content)
        elif request.is_json:
            payload = request.get_json()
        else:
            return jsonify({'success': False, 'message': 'No backup file or JSON payload provided'}), 400

        res = import_user_data_payload(current_user, payload)
        if res:
            flash('Successfully restored planner data from local JSON backup across all tables!', 'success')
            return jsonify({'success': True, 'message': 'Local backup restored successfully across all tables!'})
        return jsonify({'success': False, 'message': 'Failed to restore local backup'}), 400
    except json.JSONDecodeError:
        return jsonify({'success': False, 'message': 'Invalid JSON file format'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Restore error: {str(e)}'}), 500



# Google Drive Folder Listing Endpoint
@planner.route('/api/google/drive/folders', methods=['GET'])
@login_required
def api_google_drive_folders():
    from app.services.google_service import list_google_drive_folders
    try:
        parent_id = request.args.get('parent_id', 'root')
        folders = list_google_drive_folders(current_user, parent_id=parent_id)
        selected_folder_id = current_user.google_drive_folder_id or 'root'
        selected_folder_name = current_user.google_drive_folder_name or 'My Drive (Root Folder)'
        selected_folder_path = current_user.google_drive_folder_path or selected_folder_name
        return jsonify({
            'success': True,
            'folders': folders,
            'parent_id': parent_id,
            'current_folder_id': selected_folder_id,
            'current_folder_name': selected_folder_name,
            'current_folder_path': selected_folder_path
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Folder fetch error: {str(e)}'}), 500


# Google Drive Folder Settings Endpoint
@planner.route('/api/google/drive/folder_settings', methods=['POST'])
@login_required
def api_google_drive_folder_settings():
    from app.services.google_service import create_google_drive_folder
    data = request.get_json() or {}
    folder_action = data.get('action', 'select')  # 'select' or 'create'
    parent_id = data.get('parent_id', 'root')
    folder_path = data.get('folder_path', '')

    if folder_action == 'create':
        folder_name = data.get('folder_name', 'chronos planner folder')
        folder_info = create_google_drive_folder(current_user, folder_name, parent_id=parent_id)
        folder_id = folder_info['id']
        folder_name = folder_info['name']
    else:
        folder_id = data.get('folder_id', 'root')
        folder_name = data.get('folder_name', 'My Drive (Root Folder)')

    if not folder_path:
        folder_path = folder_name

    current_user.google_drive_folder_id = folder_id
    current_user.google_drive_folder_name = folder_name
    current_user.google_drive_folder_path = folder_path
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Backup target folder set to "{folder_name}"',
        'folder_id': folder_id,
        'folder_name': folder_name,
        'folder_path': folder_path
    })




# Excel Export Helper
def create_styled_excel(title, sheets_data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)  # Remove default active sheet

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for sheet_title, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_title[:31])
        ws.cell(row=1, column=1, value=title).font = title_font
        ws.append([])

        if rows:
            header_row = rows[0]
            ws.append(header_row)
            header_idx = ws.max_row
            for col in range(1, len(header_row) + 1):
                cell = ws.cell(row=header_idx, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in rows[1:]:
                ws.append(row)
                row_idx = ws.max_row
                for col in range(1, len(row) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = cell_font
                    cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# 1. Daily Planner Excel Export
@planner.route('/daily/export_excel')
@login_required
def export_daily_excel():
    date_str = request.args.get('date')
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = get_today_date()
    else:
        target_date = get_today_date()

    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=target_date).first()

    # Tasks Sheet
    tasks_rows = [["Task ID", "Task Description", "Priority Level", "Completion Status"]]
    if plan and plan.tasks:
        for t in plan.tasks:
            tasks_rows.append([
                str(t.get('id', '')),
                str(t.get('text', '')),
                str(t.get('priority', 'Medium')),
                "Completed" if t.get('completed') else "Pending"
            ])

    # Hourly Schedule & Mood Sheet
    default_slots = [
        "05:00 - 06:00 AM", "06:00 - 07:00 AM", "07:00 - 08:00 AM", "08:00 - 09:00 AM",
        "09:00 - 10:00 AM", "10:00 - 11:00 AM", "11:00 - 12:00 PM", "12:00 - 01:00 PM",
        "01:00 - 02:00 PM", "02:00 - 03:00 PM", "03:00 - 04:00 PM", "04:00 - 05:00 PM",
        "05:00 - 06:00 PM", "06:00 - 07:00 PM", "07:00 - 08:00 PM", "08:00 - 09:00 PM",
        "09:00 - 10:00 PM", "10:00 - 11:00 PM", "11:00 - 12:00 AM"
    ]
    schedule_rows = [["Time Slot", "Activity Details", "Hourly Mood Tag"]]
    raw_schedule = plan.schedule if plan and plan.schedule else {}
    for slot in default_slots:
        val = raw_schedule.get(slot, {})
        act = val.get('activity', '') if isinstance(val, dict) else str(val or '')
        mood = val.get('mood', '') if isinstance(val, dict) else ''
        schedule_rows.append([slot, act, mood])

    # Depression Episodes Sheet
    episodes_rows = [["Episode ID", "Time Logged", "Start Time", "Duration", "Intensity (1-10)", "Triggers / Symptoms", "Coping Mechanism", "Effectiveness", "Personal Notes"]]
    if plan and plan.depression_episodes:
        for ep in plan.depression_episodes:
            episodes_rows.append([
                str(ep.get('id', '')),
                str(ep.get('entry_time', '')),
                str(ep.get('start_time', '')),
                str(ep.get('duration', '')),
                str(ep.get('intensity', 5)),
                str(ep.get('triggers', '')),
                str(ep.get('coping_mechanism', '')),
                str(ep.get('coping_effectiveness', '')),
                str(ep.get('notes', ''))
            ])

    # Memory Tracker Sheet
    memory_rows = [["Log ID", "Time Logged", "Time of Slip", "Forgotten Detail / Task", "Category", "Trigger / Context", "Impact Level", "Recovery Status", "Notes"]]
    if plan and plan.memory_logs:
        for m in plan.memory_logs:
            memory_rows.append([
                str(m.get('id', '')),
                str(m.get('entry_time', '')),
                str(m.get('time', '')),
                str(m.get('item', '')),
                str(m.get('category', '')),
                str(m.get('context', '')),
                str(m.get('impact', '')),
                str(m.get('recovery', '')),
                str(m.get('notes', ''))
            ])

    # Sleep Tracker Sheet
    s_log = plan.sleep_log if plan and plan.sleep_log else {}
    sleep_rows = [
        ["Metric", "Value"],
        ["Sleep Duration (Hours)", f"{s_log.get('hours', 0)} hrs"],
        ["Bedtime", s_log.get('bedtime', 'N/A')],
        ["Wake-up Time", s_log.get('wake_time', 'N/A')],
        ["Sleep Quality (1-10)", s_log.get('quality', 'N/A')],
        ["Disruptions", s_log.get('disruptions', 'None')],
        ["Sleep Notes & Factors", s_log.get('notes', '')]
    ]

    # Notes Sheet
    notes_rows = [["Section", "Content"], ["Daily Reflection & Notes", plan.notes if plan else '']]

    excel_file = create_styled_excel(
        f"Chronos Daily Planner - {target_date.strftime('%B %d, %Y')}",
        {
            "Daily Tasks": tasks_rows,
            "Hourly Activity & Mood": schedule_rows,
            "Sleep Tracker": sleep_rows,
            "Depression Tracker": episodes_rows,
            "Memory Tracker": memory_rows,
            "Daily Reflection": notes_rows
        }
    )

    filename = f"Daily_Planner_{target_date.strftime('%Y_%m_%d')}.xlsx"
    return send_file(excel_file, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 2. Weekly Planner Excel Export
@planner.route('/weekly/export_excel')
@login_required
def export_weekly_excel():
    year_param = request.args.get('year', type=int)
    week_param = request.args.get('week', type=int)
    today = get_today_date()
    if not year_param or not week_param:
        year_param, week_param, _ = today.isocalendar()

    plan = WeeklyPlan.query.filter_by(user_id=current_user.id, year=year_param, week_number=week_param).first()

    # Goals Sheet
    goals_rows = [["Goal ID", "Weekly Goal Title", "Status"]]
    if plan and plan.goals:
        for g in plan.goals:
            goals_rows.append([str(g.get('id', '')), str(g.get('title', '')), "Achieved" if g.get('completed') else "In Progress"])

    # Shopping List Sheet
    shopping_rows = [["Item ID", "Shopping Item Name", "Category", "Bought Status"]]
    if plan and plan.shopping_list:
        for s in plan.shopping_list:
            shopping_rows.append([str(s.get('id', '')), str(s.get('item', '')), str(s.get('category', 'Groceries')), "Bought" if s.get('bought') else "Pending"])

    # 7-Day To-Dos Grid Sheet
    day_abbrs = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    todos_rows = [["Day", "To-Do Description", "Status"]]
    daily_todos = plan.daily_todos if plan and plan.daily_todos else {}
    for d in day_abbrs:
        for t in daily_todos.get(d, []):
            todos_rows.append([d, str(t.get('text', '')), "Completed" if t.get('completed') else "Pending"])

    # Meals Menu Sheet
    meals_rows = [["Day", "Breakfast", "Lunch", "Dinner"]]
    meals_menu = plan.meals_menu if plan and plan.meals_menu else {}
    for d in day_abbrs:
        dm = meals_menu.get(d, {})
        meals_rows.append([d, str(dm.get('breakfast', '')), str(dm.get('lunch', '')), str(dm.get('dinner', ''))])

    # Notes Sheet
    notes_rows = [["Section", "Content"], ["Weekly Reflection & Notes", plan.notes if plan else '']]

    excel_file = create_styled_excel(
        f"Chronos Weekly Planner - Week {week_param}, {year_param}",
        {
            "Weekly Goals": goals_rows,
            "Shopping List": shopping_rows,
            "7-Day To-Dos": todos_rows,
            "Meals Menu": meals_rows,
            "Weekly Reflection": notes_rows
        }
    )

    filename = f"Weekly_Planner_{year_param}_W{week_param:02d}.xlsx"
    return send_file(excel_file, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 3. Monthly Planner Excel Export
@planner.route('/monthly/export_excel')
@login_required
def export_monthly_excel():
    today = get_today_date()
    selected_year = request.args.get('year', type=int, default=today.year)
    selected_month = request.args.get('month', type=int, default=today.month)
    month_name = calendar.month_name[selected_month]

    plan = MonthlyPlan.query.filter_by(user_id=current_user.id, year=selected_year, month=selected_month).first()

    # Goals Sheet
    goals_rows = [["Goal ID", "Monthly Goal Title", "Category", "Status"]]
    if plan and plan.goals:
        for g in plan.goals:
            goals_rows.append([str(g.get('id', '')), str(g.get('title', '')), str(g.get('category', 'Personal')), str(g.get('status', 'In Progress'))])

    # Milestones Sheet
    milestones_rows = [["Milestone ID", "Milestone Description", "Target Day", "Status"]]
    if plan and plan.milestones:
        for m in plan.milestones:
            milestones_rows.append([str(m.get('id', '')), str(m.get('title', '')), str(m.get('day', '')), "Achieved" if m.get('completed') else "Pending"])

    # Habits Sheet
    habits_rows = [["Habit ID", "Habit Name", "Days Completed Count"]]
    if plan and plan.habits:
        for h in plan.habits:
            habits_rows.append([str(h.get('id', '')), str(h.get('name', '')), str(len(h.get('completed_days', [])))])

    # Notes Sheet
    notes_rows = [["Section", "Content"], ["Monthly Reflections & Notes", plan.notes if plan else '']]

    excel_file = create_styled_excel(
        f"Chronos Monthly Planner - {month_name} {selected_year}",
        {
            "Monthly Goals": goals_rows,
            "Key Milestones": milestones_rows,
            "Tracked Habits": habits_rows,
            "Monthly Notes": notes_rows
        }
    )

    filename = f"Monthly_Planner_{selected_year}_{selected_month:02d}.xlsx"
    return send_file(excel_file, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# 4. Yearly Planner Excel Export
@planner.route('/yearly/export_excel')
@login_required
def export_yearly_excel():
    today = get_today_date()
    selected_year = request.args.get('year', type=int, default=today.year)

    plan = YearlyPlan.query.filter_by(user_id=current_user.id, year=selected_year).first()

    # Resolutions Sheet
    resolutions_rows = [["Resolution ID", "Resolution Text", "Category", "Status"]]
    if plan and plan.resolutions:
        for r in plan.resolutions:
            resolutions_rows.append([str(r.get('id', '')), str(r.get('text', '')), str(r.get('category', 'Personal')), "Accomplished" if r.get('completed') else "In Progress"])

    # Objectives Sheet
    objectives_rows = [["Objective ID", "Strategic Objective Title", "Quarter", "Status"]]
    if plan and plan.objectives:
        for o in plan.objectives:
            objectives_rows.append([str(o.get('id', '')), str(o.get('title', '')), str(o.get('quarter', 'Q1')), str(o.get('status', 'In Progress'))])

    # 12-Month Achievement Grid Sheet
    all_monthly_plans = MonthlyPlan.query.filter_by(user_id=current_user.id, year=selected_year).all()
    monthly_map = {mp.month: mp for mp in all_monthly_plans}
    grid_rows = [["Month #", "Month Name", "Goals Accomplished", "Milestones Reached", "Habits Tracked", "Achievement Score"]]
    for m in range(1, 13):
        m_name = calendar.month_name[m]
        mp = monthly_map.get(m)
        m_goals = mp.goals if mp and mp.goals else []
        m_goals_completed = sum(1 for g in m_goals if g.get('status') == 'Completed')
        m_milestones = mp.milestones if mp and mp.milestones else []
        m_milestones_completed = sum(1 for ms in m_milestones if ms.get('completed'))
        m_habits = mp.habits if mp and mp.habits else []
        
        comps = []
        if len(m_goals) > 0:
            comps.append(int(m_goals_completed / len(m_goals) * 100))
        if len(m_milestones) > 0:
            comps.append(int(m_milestones_completed / len(m_milestones) * 100))
        score = int(sum(comps) / len(comps)) if comps else (100 if m_goals_completed > 0 or m_milestones_completed > 0 else 0)

        grid_rows.append([str(m), m_name, f"{m_goals_completed}/{len(m_goals)}", f"{m_milestones_completed}/{len(m_milestones)}", str(len(m_habits)), f"{score}%"])

    # Reflections Sheet
    reflections_rows = [["Section", "Content"], ["Year-in-Review Reflections", plan.reflections if plan else '']]

    excel_file = create_styled_excel(
        f"Chronos Yearly Planner - Year {selected_year}",
        {
            "Annual Resolutions": resolutions_rows,
            "Strategic Objectives": objectives_rows,
            "12-Month Achievement Grid": grid_rows,
            "Year Reflections": reflections_rows
        }
    )

    filename = f"Yearly_Planner_{selected_year}.xlsx"
    return send_file(excel_file, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# PLANNING TAB — Persistent, date-independent task checklist
# ---------------------------------------------------------------------------

@planner.route('/planning')
@login_required
def planning():
    """Render the Planning page with all persistent tasks for the current user."""
    user_tags = get_user_tags(current_user)
    tasks = (
        PlanningTask.query
        .filter_by(user_id=current_user.id)
        .order_by(PlanningTask.sort_order.asc(), PlanningTask.created_at.asc())
        .all()
    )
    return render_template(
        'planner/planning.html',
        tasks=tasks,
        user_tags=user_tags,
        today=get_today_date()
    )


@planner.route('/api/planning/task/add', methods=['POST'])
@login_required
def api_planning_add_task():
    data = request.get_json() or {}
    text = data.get('text', '').strip()
    priority = data.get('priority', 'Medium')
    tags = data.get('tags', [])
    if isinstance(tags, str):
        tags = [x.strip() for x in tags.split(',') if x.strip()]

    if not text:
        return jsonify({'success': False, 'message': 'Task text is required'}), 400

    max_order = db.session.query(db.func.max(PlanningTask.sort_order)).filter_by(user_id=current_user.id).scalar() or 0

    task = PlanningTask(
        user_id=current_user.id,
        text=text,
        priority=priority,
        tags=tags,
        completed=False,
        sort_order=max_order + 1
    )
    db.session.add(task)
    db.session.commit()
    return jsonify({'success': True, 'task': task.to_dict()})


@planner.route('/api/planning/task/toggle', methods=['POST'])
@login_required
def api_planning_toggle_task():
    data = request.get_json() or {}
    task_id = data.get('task_id')
    if not task_id:
        return jsonify({'success': False, 'message': 'Missing task_id'}), 400

    task = PlanningTask.query.filter_by(id=task_id, user_id=current_user.id).first()
    if not task:
        return jsonify({'success': False, 'message': 'Task not found'}), 404

    task.completed = not task.completed
    db.session.commit()
    return jsonify({'success': True, 'completed': task.completed})


@planner.route('/api/planning/task/edit', methods=['POST'])
@login_required
def api_planning_edit_task():
    data = request.get_json() or {}
    task_id = data.get('task_id')
    if not task_id:
        return jsonify({'success': False, 'message': 'Missing task_id'}), 400

    task = PlanningTask.query.filter_by(id=task_id, user_id=current_user.id).first()
    if not task:
        return jsonify({'success': False, 'message': 'Task not found'}), 404

    if 'text' in data and data['text'].strip():
        task.text = data['text'].strip()
    if 'priority' in data:
        task.priority = data['priority']
    if 'tags' in data:
        tags = data['tags']
        if isinstance(tags, str):
            tags = [x.strip() for x in tags.split(',') if x.strip()]
        task.tags = tags

    db.session.commit()
    return jsonify({'success': True, 'task': task.to_dict()})


@planner.route('/api/planning/task/delete', methods=['POST'])
@login_required
def api_planning_delete_task():
    data = request.get_json() or {}
    task_id = data.get('task_id')
    if not task_id:
        return jsonify({'success': False, 'message': 'Missing task_id'}), 400

    task = PlanningTask.query.filter_by(id=task_id, user_id=current_user.id).first()
    if not task:
        return jsonify({'success': False, 'message': 'Task not found'}), 404

    db.session.delete(task)
    db.session.commit()
    return jsonify({'success': True})


@planner.route('/api/planning/task/move_to_daily', methods=['POST'])
@login_required
def api_planning_move_to_daily():
    """Copy a planning task into today's DailyPlan tasks list, then delete from Planning."""
    data = request.get_json() or {}
    task_id = data.get('task_id')
    if not task_id:
        return jsonify({'success': False, 'message': 'Missing task_id'}), 400

    task = PlanningTask.query.filter_by(id=task_id, user_id=current_user.id).first()
    if not task:
        return jsonify({'success': False, 'message': 'Task not found'}), 404

    today = get_today_date()
    plan = DailyPlan.query.filter_by(user_id=current_user.id, date=today).first()
    if not plan:
        plan = DailyPlan(user_id=current_user.id, date=today, schedule={}, tasks=[], notes='')
        db.session.add(plan)

    daily_tasks = list(plan.tasks or [])
    new_daily_task = {
        'id': str(int(datetime.utcnow().timestamp() * 1000)),
        'text': task.text,
        'priority': task.priority,
        'tags': task.tags or [],
        'status': 'To Do',
        'note': '',
        'completed': False,
        'is_default': False,
        'is_spillover': False,
        'spillover_count': 0,
        'original_date': today.strftime('%Y-%m-%d')
    }
    daily_tasks.append(new_daily_task)
    plan.tasks = daily_tasks
    flag_modified(plan, 'tasks')

    db.session.delete(task)
    db.session.commit()
    return jsonify({'success': True, 'message': f"Task moved to today's Daily checklist ({today.strftime('%b %d')})"})


@planner.route('/api/planning/task/reorder', methods=['POST'])
@login_required
def api_planning_reorder_tasks():
    data = request.get_json() or {}
    task_ids = data.get('task_ids', [])
    if not isinstance(task_ids, list):
        return jsonify({'success': False, 'message': 'task_ids must be a list'}), 400

    tasks = PlanningTask.query.filter_by(user_id=current_user.id).all()
    task_map = {t.id: t for t in tasks}

    for idx, tid in enumerate(task_ids):
        try:
            tid_int = int(tid)
        except (ValueError, TypeError):
            continue
        if tid_int in task_map:
            task_map[tid_int].sort_order = idx

    db.session.commit()
    return jsonify({'success': True})
