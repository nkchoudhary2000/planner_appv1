import os
import io
import calendar
from datetime import date, datetime, timedelta, timezone
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

from flask import g
from flask_login import current_user
from sqlalchemy.orm.attributes import flag_modified
from app import db
from app.models import DailyPlan, MonthlyPlan, YearlyPlan, WeeklyPlan, User, PlanningTask

DEFAULT_TAGS = [
    {'id': 'tag_work', 'name': 'Work', 'color': '#3b82f6'},
    {'id': 'tag_personal', 'name': 'Personal', 'color': '#10b981'},
    {'id': 'tag_health', 'name': 'Health', 'color': '#ec4899'},
    {'id': 'tag_finance', 'name': 'Finance', 'color': '#f59e0b'},
    {'id': 'tag_urgent', 'name': 'Urgent', 'color': '#ef4444'}
]


def get_current_user_safe():
    """Retrieve currently authenticated user from g.current_user or flask_login current_user."""
    user = getattr(g, 'current_user', None)
    if user:
        return user
    if current_user and current_user.is_authenticated:
        return current_user
    return None


def get_user_tags(user):
    """Return configured custom tags for the user or DEFAULT_TAGS."""
    if user and hasattr(user, 'custom_tags') and user.custom_tags:
        return user.custom_tags
    return DEFAULT_TAGS


def get_today_date():
    """Return current date in configured timezone (default Asia/Kolkata / IST UTC+5:30)."""
    tz_name = os.environ.get('APP_TIMEZONE', 'Asia/Kolkata')
    if ZoneInfo:
        try:
            return datetime.now(ZoneInfo(tz_name)).date()
        except Exception:
            pass
    return datetime.now(timezone(timedelta(hours=5, minutes=30))).date()


def to_24h_time(time_str, default="23:00"):
    if not time_str or not isinstance(time_str, str):
        return default
    time_str = time_str.strip()
    for fmt in ('%I:%M %p', '%I:%M%p', '%H:%M', '%H:%M:%S'):
        try:
            dt = datetime.strptime(time_str, fmt)
            return dt.strftime('%H:%M')
        except ValueError:
            pass
    return default


def format_time_12h(time_str, default="11:00 PM"):
    if not time_str or not isinstance(time_str, str):
        return default
    time_str = time_str.strip()
    for fmt in ('%H:%M', '%I:%M %p', '%I:%M%p', '%H:%M:%S'):
        try:
            dt = datetime.strptime(time_str, fmt)
            return dt.strftime('%I:%M %p').lstrip('0')
        except ValueError:
            pass
    return time_str


def process_task_spillovers(user_id, target_date):
    """
    Automatically rolls over uncompleted non-default tasks from past dates up to target_date.
    Increments spillover_count (pending days) and escalates priority up to 'High' (max severity).
    Restricted to past 30 days for maximum performance.
    """
    cutoff_date = target_date - timedelta(days=30)
    past_plans = DailyPlan.query.filter(
        DailyPlan.user_id == user_id,
        DailyPlan.date < target_date,
        DailyPlan.date >= cutoff_date
    ).order_by(DailyPlan.date.asc()).all()

    if not past_plans:
        return

    # Quick check: if no past plan contains uncompleted non-default tasks, return immediately!
    has_uncompleted = any(
        any(isinstance(t, dict) and not t.get('completed') and not t.get('is_default') for t in (p.tasks or []))
        for p in past_plans
    )
    if not has_uncompleted:
        return

    plan_map = {p.date: p for p in past_plans}
    earliest_date = past_plans[0].date

    curr_date = earliest_date
    while curr_date < target_date:
        next_date = curr_date + timedelta(days=1)
        curr_plan = plan_map.get(curr_date)

        if curr_plan and curr_plan.tasks:
            # Only non-default uncompleted tasks spill over
            uncompleted = [t for t in curr_plan.tasks if isinstance(t, dict) and not t.get('completed') and not t.get('is_default')]
            if uncompleted:
                next_plan = plan_map.get(next_date)
                if not next_plan:
                    next_plan = DailyPlan.query.filter_by(user_id=user_id, date=next_date).first()
                if not next_plan:
                    next_plan = DailyPlan(user_id=user_id, date=next_date, schedule={}, tasks=[], notes='')
                    db.session.add(next_plan)

                next_tasks = list(next_plan.tasks or [])
                existing_ids = {t.get('id') for t in next_tasks if isinstance(t, dict) and t.get('id')}

                dismissed_ids = set()
                raw_sched = next_plan.schedule or {}
                if isinstance(raw_sched, dict):
                    dismissed_ids = set(raw_sched.get('_dismissed_tasks', []))

                modified_next = False

                for t in uncompleted:
                    t_id = t.get('id')
                    if t_id and t_id not in existing_ids and t_id not in dismissed_ids:
                        old_priority = t.get('priority', 'Medium')
                        if old_priority == 'Low':
                            new_priority = 'Medium'
                        else:
                            new_priority = 'High'  # Escalates to High (max severity)

                        spill_count = t.get('spillover_count', 0) + 1
                        orig_date = t.get('original_date', curr_date.strftime('%Y-%m-%d'))

                        spill_task = {
                            'id': t_id,
                            'text': t.get('text', ''),
                            'priority': new_priority,
                            'completed': False,
                            'is_spillover': True,
                            'spillover_count': spill_count,
                            'original_date': orig_date
                        }
                        next_tasks.append(spill_task)
                        existing_ids.add(t_id)
                        modified_next = True

                if modified_next:
                    next_plan.tasks = next_tasks
                    flag_modified(next_plan, 'tasks')
                    plan_map[next_date] = next_plan

        curr_date = next_date

    db.session.commit()


def populate_daily_defaults(user_id, target_date):
    """
    Populates default daily routine tasks and default activity schedule slots
    for target_date based on the latest saved default configurations.
    """
    past_plans = DailyPlan.query.filter(
        DailyPlan.user_id == user_id,
        DailyPlan.date <= target_date
    ).order_by(DailyPlan.date.desc()).limit(30).all()

    if not past_plans:
        return

    # Resolve active Default Tasks based on the MOST RECENT plan setting/action for each task
    default_tasks_master = {}
    seen_task_keys = set()

    for p in past_plans:
        raw_sched = p.schedule or {}
        deleted_on_plan = set()
        if isinstance(raw_sched, dict):
            for dkey in raw_sched.get('_deleted_defaults', []):
                deleted_on_plan.add(str(dkey).lower())

        if p.tasks:
            for t in p.tasks:
                if isinstance(t, dict):
                    t_text = t.get('text', '').strip()
                    t_id = str(t.get('id', '')).lower()
                    t_key = t_text.lower()

                    if t_key and t_key not in seen_task_keys:
                        seen_task_keys.add(t_key)
                        # If marked default and not deleted on this same plan
                        if t.get('is_default') and t_key not in deleted_on_plan and t_id not in deleted_on_plan:
                            default_tasks_master[t_key] = {
                                'text': t_text,
                                'priority': t.get('priority', 'Medium'),
                                'is_default': True
                            }

        # Any default deletion recorded on this plan that wasn't already determined by a newer plan
        for dkey in deleted_on_plan:
            seen_task_keys.add(dkey)

    # Resolve Default Schedule Slots based on the MOST RECENT plan setting for each slot
    default_schedule_master = {}
    resolved_slots = set()

    for p in past_plans:
        if p.schedule:
            for slot, sdata in p.schedule.items():
                if slot.startswith('_') or slot in resolved_slots:
                    continue

                if isinstance(sdata, dict):
                    if p.date < target_date or sdata.get('activity') or sdata.get('is_default'):
                        resolved_slots.add(slot)
                        if sdata.get('is_default') and sdata.get('activity'):
                            default_schedule_master[slot] = {
                                'activity': sdata.get('activity', ''),
                                'mood': sdata.get('mood', ''),
                                'is_default': True
                            }

    if not default_tasks_master and not default_schedule_master:
        return

    plan = DailyPlan.query.filter_by(user_id=user_id, date=target_date).first()
    if not plan:
        plan = DailyPlan(user_id=user_id, date=target_date, schedule={}, tasks=[], notes='')
        db.session.add(plan)

    tasks = list(plan.tasks or [])
    schedule = dict(plan.schedule or {})
    modified = False

    target_dismissed = set()
    if isinstance(schedule, dict):
        target_dismissed = set(schedule.get('_dismissed_tasks', []))
        target_dismissed.update(schedule.get('_deleted_defaults', []))

    target_dismissed_lower = {str(x).lower() for x in target_dismissed}

    # Populate missing default tasks
    existing_texts = {t.get('text', '').strip().lower() for t in tasks if isinstance(t, dict) and t.get('text')}
    for key_lower, dtask in default_tasks_master.items():
        if key_lower not in existing_texts and key_lower not in target_dismissed_lower:
            new_id = f"def_{int(datetime.utcnow().timestamp() * 1000)}_{len(tasks)}"
            if new_id.lower() not in target_dismissed_lower:
                tasks.append({
                    'id': new_id,
                    'text': dtask['text'],
                    'priority': dtask['priority'],
                    'completed': False,
                    'is_default': True,
                    'is_spillover': False,
                    'spillover_count': 0,
                    'original_date': target_date.strftime('%Y-%m-%d')
                })
                modified = True

    # Populate missing default schedule slots
    for slot, dslot in default_schedule_master.items():
        curr_val = schedule.get(slot)
        is_empty = False
        if not curr_val:
            is_empty = True
        elif isinstance(curr_val, dict) and not curr_val.get('activity'):
            is_empty = True

        if is_empty:
            schedule[slot] = {
                'activity': dslot['activity'],
                'mood': dslot['mood'],
                'is_default': True
            }
            modified = True

    if modified:
        plan.tasks = tasks
        plan.schedule = schedule
        flag_modified(plan, 'tasks')
        flag_modified(plan, 'schedule')
        db.session.commit()


def carry_forward_unbought_shopping_items(user_id, target_year, target_week):
    """
    Moves unbought shopping list items (bought: False) from prior weekly plans
    to the target week's shopping list, and removes them from prior weekly plans.
    """
    first_day_of_year = date(target_year, 1, 4)
    start_of_target_week = first_day_of_year + timedelta(weeks=target_week - 1) - timedelta(days=first_day_of_year.weekday())

    target_plan = WeeklyPlan.query.filter_by(
        user_id=user_id,
        year=target_year,
        week_number=target_week
    ).first()

    prior_plans = WeeklyPlan.query.filter(
        WeeklyPlan.user_id == user_id,
        WeeklyPlan.start_date < start_of_target_week
    ).order_by(WeeklyPlan.start_date.asc()).all()

    items_to_move = []
    modified_priors = False

    for prior_plan in prior_plans:
        p_list = prior_plan.shopping_list or []
        if not p_list:
            continue
        
        bought_items = []
        unbought_items = []

        for item in p_list:
            if isinstance(item, dict):
                if item.get('bought'):
                    bought_items.append(item)
                else:
                    unbought_items.append(item)

        if unbought_items:
            items_to_move.extend(unbought_items)
            prior_plan.shopping_list = bought_items
            flag_modified(prior_plan, 'shopping_list')
            modified_priors = True

    if items_to_move:
        if not target_plan:
            target_plan = WeeklyPlan(
                user_id=user_id,
                year=target_year,
                week_number=target_week,
                start_date=start_of_target_week,
                goals=[],
                daily_todos={abbr: [] for abbr in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']},
                shopping_list=[],
                meals_menu={abbr: {'breakfast': '', 'lunch': '', 'dinner': ''} for abbr in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']},
                notes=''
            )
            db.session.add(target_plan)

        target_list = list(target_plan.shopping_list or [])
        existing_ids = {s.get('id') for s in target_list if isinstance(s, dict) and s.get('id')}

        for item in items_to_move:
            if item.get('id') not in existing_ids:
                target_list.append(item)

        target_plan.shopping_list = target_list
        flag_modified(target_plan, 'shopping_list')
        db.session.commit()
    elif modified_priors:
        db.session.commit()

    return target_plan


def create_styled_excel(title, sheets_data):
    """Generate a styled Excel spreadsheet from a dictionary of sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)  # Remove default active sheet

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for sheet_title, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_title[:31])
        ws.cell(row=1, column=1, value=title).font = title_font
        ws.append([])

        if rows:
            header_row = rows[0]
            ws.append(header_row)
            header_idx = ws.max_row
            for col in range(1, len(header_row) + 1):
                cell = ws.cell(row=header_idx, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in rows[1:]:
                ws.append(row)
                row_idx = ws.max_row
                for col in range(1, len(row) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = cell_font
                    cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def copy_habits_from_previous_month(user_id, target_year, target_month):
    """
    Copies habits from the previous month into the target month's MonthlyPlan.
    - Generates fresh habit entries with empty completions for the target month.
    - If a habit name already exists in target month, renames duplicate with ' (Copy)' marker.
    - Keeps previous month's plan, completions, and logs completely undisturbed.
    Returns dict: {'success': bool, 'message': str, 'imported_count': int, 'habits': list, 'target_plan': MonthlyPlan}
    """
    if target_month == 1:
        prev_year, prev_month = target_year - 1, 12
    else:
        prev_year, prev_month = target_year, target_month - 1

    prev_month_name = calendar.month_name[prev_month]

    prev_plan = MonthlyPlan.query.filter_by(
        user_id=user_id, year=prev_year, month=prev_month
    ).first()

    if not prev_plan or not prev_plan.habits:
        return {
            'success': False,
            'message': f'No habits found in previous month ({prev_month_name} {prev_year}).',
            'imported_count': 0,
            'habits': []
        }

    valid_prev_habits = [h for h in prev_plan.habits if (h.get('name') or '').strip()]
    if not valid_prev_habits:
        return {
            'success': False,
            'message': f'No valid habits found in previous month ({prev_month_name} {prev_year}).',
            'imported_count': 0,
            'habits': []
        }

    plan = MonthlyPlan.query.filter_by(
        user_id=user_id, year=target_year, month=target_month
    ).first()

    if not plan:
        plan = MonthlyPlan(
            user_id=user_id,
            year=target_year,
            month=target_month,
            goals=[],
            habits=[],
            milestones=[],
            calendar_days={},
            notes=''
        )
        db.session.add(plan)

    current_habits = list(plan.habits or [])
    existing_names_lower = [h.get('name', '').strip().lower() for h in current_habits if h.get('name')]

    new_habits_added = []
    base_ts = int(datetime.utcnow().timestamp() * 1000)

    for idx, prev_h in enumerate(valid_prev_habits):
        orig_name = prev_h.get('name', '').strip()
        
        # Check if already exists in target month -> append '(Copy)' suffix
        if orig_name.lower() in existing_names_lower:
            candidate_name = f"{orig_name} (Copy)"
            copy_counter = 2
            while candidate_name.lower() in existing_names_lower:
                candidate_name = f"{orig_name} (Copy {copy_counter})"
                copy_counter += 1
            final_name = candidate_name
        else:
            final_name = orig_name

        existing_names_lower.append(final_name.lower())

        h_type = (prev_h.get('type') or 'boolean').strip().lower()
        if h_type not in ['boolean', 'counter', 'sub_habits']:
            h_type = 'boolean'

        new_h = {
            'id': f"{base_ts}_{len(current_habits) + 1 + idx}",
            'name': final_name,
            'type': h_type,
            'category': prev_h.get('category', 'General'),
            'completed_days': []  # Fresh slate for target month
        }

        if prev_h.get('is_github'):
            new_h['is_github'] = True

        if h_type == 'counter':
            new_h['unit'] = prev_h.get('unit', 'times')
            new_h['target_count'] = max(1, int(prev_h.get('target_count', 1)))
            new_h['daily_counts'] = {}  # Fresh slate
        elif h_type == 'sub_habits':
            sub_list = []
            for s_idx, s in enumerate(prev_h.get('sub_habits', [])):
                if isinstance(s, dict):
                    s_name = (s.get('name') or '').strip()
                else:
                    s_name = str(s).strip()
                if s_name:
                    sub_list.append({
                        'id': f"sh_{base_ts}_{idx}_{s_idx + 1}",
                        'name': s_name
                    })
            new_h['sub_habits'] = sub_list
            new_h['daily_sub_completions'] = {}  # Fresh slate

        current_habits.append(new_h)
        new_habits_added.append(new_h)

    plan.habits = current_habits
    flag_modified(plan, 'habits')
    db.session.commit()

    return {
        'success': True,
        'message': f'Successfully imported {len(new_habits_added)} habit(s) from {prev_month_name} {prev_year}!',
        'imported_count': len(new_habits_added),
        'habits': plan.habits
    }

